"""写真の抽出・リサイズ・Excel挿入モジュール

ラインナップExcelからLED商品写真を抽出し、
見積テンプレートの各シートに挿入する機能を提供する。
"""

from __future__ import annotations

import io
import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional, Union

import openpyxl
from openpyxl.drawing.image import Image as XlImage
from openpyxl.drawing.spreadsheet_drawing import (
    AnchorMarker,
    OneCellAnchor,
)
from openpyxl.utils.units import pixels_to_EMU
from PIL import Image as PILImage

from models import LEDProduct

logger = logging.getLogger(__name__)

# ===== 写真サイズ定数 =====
# ⑩内訳シートの既存/LED写真
BREAKDOWN_PHOTO_W = 100
BREAKDOWN_PHOTO_H = 90

# ⑪除外シートの除外器具写真
EXCLUSION_PHOTO_W = 78
EXCLUSION_PHOTO_H = 80

# 選定シートのLED商品写真
SELECTION_PHOTO1_W = 46   # A列 (写真①)
SELECTION_PHOTO2_W = 58   # B列 (写真②)
SELECTION_PHOTO_H = 88


@dataclass
class ImageRef:
    """ラインナップExcel内の画像参照"""
    source_file: str
    source_sheet: str
    anchor_row: int     # 0-based row
    anchor_col: int     # 0=A, 1=B
    img_index: int      # ws._images 内のインデックス


class LineupImageIndex:
    """ラインナップExcelの商品画像をインデックス化

    (source_file, source_sheet, anchor_row_1based) → ImageRef のマッピング。
    画像データは必要時にExcelファイルから遅延取得する。
    """

    def __init__(self):
        # key: (file, sheet, row_1based, col) → ImageRef
        self._refs: dict[tuple[str, str, int, int], ImageRef] = {}
        # key: (file, sheet) → list of rows with images (sorted)
        self._rows_by_sheet: dict[tuple[str, str], list[int]] = {}
        # key: (file, sheet, row_1based, col) → bytes
        self._image_cache: dict[tuple[str, str, int, int], bytes] = {}
        self._lineup_dir: Optional[Path] = None

    def load_all(self, lineup_dir: Path) -> None:
        """ラインナップ表の全画像位置をインデックス化"""
        self._lineup_dir = lineup_dir
        from lineup_loader import LINEUP_SHEETS

        for file_enum, sheets in LINEUP_SHEETS.items():
            filepath = lineup_dir / file_enum
            if not filepath.exists():
                logger.warning(f"ファイルが見つかりません: {filepath}")
                continue

            wb = openpyxl.load_workbook(filepath, data_only=True)
            self._index_workbook(wb, file_enum, sheets)
            wb.close()

        total = len(self._refs)
        logger.info(f"画像インデックス完了: {total}件")

    def _index_workbook(
        self, wb: openpyxl.Workbook, source_file: str,
        target_sheets: list[str],
    ) -> None:
        """ワークブック内の画像位置を記録"""
        for sheet_name in target_sheets:
            # シート名が完全一致しない場合もある（末尾スペース等）
            actual_name = self._find_sheet_name(wb, sheet_name)
            if actual_name is None:
                continue

            # キーはstrip()済みのシート名を使う（lineup_loader と一致させる）
            normalized_sheet = sheet_name.strip()
            ws = wb[actual_name]
            rows_with_images = set()

            for idx, img in enumerate(ws._images):
                anchor = img.anchor
                if not hasattr(anchor, '_from'):
                    continue

                fr = anchor._from
                # A列(col=0) または B列(col=1) のみ対象
                # 大きな列番号（col=14等）は補助表示用なので除外
                if fr.col > 1:
                    continue

                row_1b = fr.row + 1  # 0-based → 1-based
                key = (source_file, normalized_sheet, row_1b, fr.col)

                if key not in self._refs:
                    self._refs[key] = ImageRef(
                        source_file=source_file,
                        source_sheet=normalized_sheet,
                        anchor_row=fr.row,
                        anchor_col=fr.col,
                        img_index=idx,
                    )
                    rows_with_images.add(row_1b)

            sheet_key = (source_file, normalized_sheet)
            self._rows_by_sheet[sheet_key] = sorted(rows_with_images)
            logger.debug(
                f"  {normalized_sheet}: {len(rows_with_images)}行に画像あり"
            )

    def _find_sheet_name(
        self, wb: openpyxl.Workbook, target: str,
    ) -> Optional[str]:
        """ワークブック内でシート名を検索（末尾スペース等に対応）"""
        for name in wb.sheetnames:
            if name.strip() == target.strip():
                return name
        return None

    def get_product_image(
        self, product: LEDProduct, photo_num: int = 1,
    ) -> Optional[bytes]:
        """LED商品の写真データを取得

        Args:
            product: LED商品
            photo_num: 1=写真①(A列), 2=写真②(B列)

        Returns:
            画像バイトデータ (PNG/JPEG)、見つからない場合None
        """
        if not self._lineup_dir:
            return None

        col = photo_num - 1  # 0=A, 1=B
        source_file = product.source_file
        source_sheet = product.source_sheet
        source_row = product.source_row  # 1-based

        # 直接マッチ
        key = (source_file, source_sheet, source_row, col)
        if key in self._refs:
            return self._extract_image(self._refs[key])

        # 上方向に最寄りの画像を探す（同一カテゴリの画像が複数行にまたがる）
        sheet_key = (source_file, source_sheet)
        rows = self._rows_by_sheet.get(sheet_key, [])

        nearest_row = None
        for r in reversed(rows):
            if r <= source_row:
                check_key = (source_file, source_sheet, r, col)
                if check_key in self._refs:
                    nearest_row = r
                    break

        if nearest_row is not None and (source_row - nearest_row) <= 8:
            # 8行以内の画像を使用
            return self._extract_image(
                self._refs[(source_file, source_sheet, nearest_row, col)]
            )

        return None

    def preload_images(self, products: list[LEDProduct]) -> None:
        """指定された商品の画像を事前にメモリにロード

        大量のget_product_image呼び出し前に一度呼ぶことで、
        ファイルの開閉回数を最小化する。
        """
        # 必要なファイル+シートを特定
        needed_files: dict[str, set[str]] = {}
        for p in products:
            if p.source_file not in needed_files:
                needed_files[p.source_file] = set()
            needed_files[p.source_file].add(p.source_sheet)

        for source_file, sheets in needed_files.items():
            filepath = self._lineup_dir / source_file
            if not filepath.exists():
                continue

            try:
                wb = openpyxl.load_workbook(filepath, data_only=True)
                for sheet_name in sheets:
                    actual_name = self._find_sheet_name(wb, sheet_name)
                    if actual_name is None:
                        continue
                    ws = wb[actual_name]
                    normalized = sheet_name.strip()

                    for img in ws._images:
                        anchor = img.anchor
                        if not hasattr(anchor, '_from'):
                            continue
                        fr = anchor._from
                        if fr.col > 1:
                            continue

                        row_1b = fr.row + 1
                        cache_key = (source_file, normalized, row_1b, fr.col)
                        if cache_key in self._refs and cache_key not in self._image_cache:
                            try:
                                self._image_cache[cache_key] = img._data()
                            except Exception:
                                pass

                wb.close()
            except Exception as e:
                logger.error(f"画像プリロードエラー ({source_file}): {e}")

        logger.info(f"画像プリロード完了: {len(self._image_cache)}枚")

    def _extract_image(self, ref: ImageRef) -> Optional[bytes]:
        """ImageRefから実際の画像バイトデータを抽出"""
        # キャッシュを確認
        cache_key = (ref.source_file, ref.source_sheet, ref.anchor_row + 1, ref.anchor_col)
        if cache_key in self._image_cache:
            return self._image_cache[cache_key]

        if not self._lineup_dir:
            return None

        filepath = self._lineup_dir / ref.source_file
        if not filepath.exists():
            return None

        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            actual_name = self._find_sheet_name(wb, ref.source_sheet)
            if actual_name is None:
                wb.close()
                return None

            ws = wb[actual_name]
            if ref.img_index < len(ws._images):
                img = ws._images[ref.img_index]
                data = img._data()
                wb.close()
                # キャッシュに保存
                self._image_cache[cache_key] = data
                return data

            wb.close()
        except Exception as e:
            logger.error(f"画像抽出エラー: {e}")

        return None

    def get_stats(self) -> dict:
        """統計情報"""
        return {
            "total_refs": len(self._refs),
            "sheets_indexed": len(self._rows_by_sheet),
        }


def _open_image(image_source: Union[Path, bytes, PILImage.Image]) -> PILImage.Image:
    """画像ソースからPIL Imageを開く"""
    if isinstance(image_source, PILImage.Image):
        img = image_source.copy()
    elif isinstance(image_source, bytes):
        img = PILImage.open(io.BytesIO(image_source))
    elif isinstance(image_source, Path):
        img = PILImage.open(image_source)
    else:
        raise ValueError(f"Unsupported image source type: {type(image_source)}")

    # RGBA→RGBに変換
    if img.mode == 'RGBA':
        bg = PILImage.new('RGB', img.size, (255, 255, 255))
        bg.paste(img, mask=img.split()[3])
        return bg
    elif img.mode != 'RGB':
        return img.convert('RGB')
    return img


# 内部データの品質倍率（表示サイズに対する実データの解像度倍率）
_QUALITY_FACTOR = 3


def resize_for_cell(
    image_source: Union[Path, bytes, PILImage.Image],
    target_width: int,
    target_height: int,
    maintain_aspect: bool = True,
) -> bytes:
    """画像をセルに合うサイズにリサイズしてJPEGバイトを返す

    内部データは表示サイズの3倍の解像度を保持し、
    Excel上での表示品質を維持する。

    Args:
        image_source: 画像パス、バイトデータ、またはPIL Image
        target_width: 目標表示幅 (px)
        target_height: 目標表示高さ (px)
        maintain_aspect: アスペクト比を維持するか

    Returns:
        高画質JPEGバイトデータ
    """
    img = _open_image(image_source)

    # 内部データは表示サイズの3倍で保持（画質維持）
    internal_w = target_width * _QUALITY_FACTOR
    internal_h = target_height * _QUALITY_FACTOR

    if maintain_aspect:
        img.thumbnail((internal_w, internal_h), PILImage.Resampling.LANCZOS)
    else:
        img = img.resize((internal_w, internal_h), PILImage.Resampling.LANCZOS)

    buf = io.BytesIO()
    img.save(buf, format='JPEG', quality=90)
    return buf.getvalue()


def prepare_fixture_photo(
    image_source: Union[Path, bytes, PILImage.Image],
    target_width: int,
    target_height: int,
    crop_ratio: float = 0.6,
) -> bytes:
    """現調写真を照明アップ用にトリミング+高画質リサイズ

    写真の中央部分を切り出して照明のクローズアップにする。
    内部データは表示サイズの3倍の解像度を保持。

    Args:
        image_source: 画像パス、バイトデータ、またはPIL Image
        target_width: 目標表示幅 (px)
        target_height: 目標表示高さ (px)
        crop_ratio: 中央トリミング比率 (0.0〜1.0)
            0.6 = 中央60%を切り出し（照明のアップ）

    Returns:
        トリミング+高画質JPEGバイトデータ
    """
    img = _open_image(image_source)
    w, h = img.size

    # 中央トリミング: crop_ratio の領域を切り出し
    crop_w = int(w * crop_ratio)
    crop_h = int(h * crop_ratio)
    left = (w - crop_w) // 2
    top = (h - crop_h) // 2
    img = img.crop((left, top, left + crop_w, top + crop_h))

    # 表示サイズの3倍で保持（画質維持）
    internal_w = target_width * _QUALITY_FACTOR
    internal_h = target_height * _QUALITY_FACTOR
    img.thumbnail((internal_w, internal_h), PILImage.Resampling.LANCZOS)

    buf = io.BytesIO()
    img.save(buf, format='JPEG', quality=90)
    return buf.getvalue()


def _get_cell_size_emu(ws, row: int, col: int) -> tuple[int, int]:
    """セルの幅・高さをEMU単位で取得

    Args:
        ws: openpyxl ワークシート
        row: 行番号 (1-based)
        col: 列番号 (1-based)

    Returns:
        (width_emu, height_emu)
    """
    from openpyxl.utils import get_column_letter

    col_letter = get_column_letter(col)
    col_width_chars = ws.column_dimensions[col_letter].width
    row_height_pts = ws.row_dimensions[row].height

    # デフォルト値
    if col_width_chars is None:
        col_width_chars = 8.43
    if row_height_pts is None:
        row_height_pts = 15.0

    # 列幅: 1文字幅 ≈ 7px → EMU変換 (1px = 9525 EMU)
    cell_w_emu = int(col_width_chars * 7 * 9525)
    # 行高: 1pt = 12700 EMU
    cell_h_emu = int(row_height_pts * 12700)

    return cell_w_emu, cell_h_emu


def insert_image_to_cell(
    ws,
    image_data: bytes,
    row: int,
    col: int,
    width_px: int,
    height_px: int,
) -> None:
    """ワークシートの指定セルに画像を中央配置で挿入

    Args:
        ws: openpyxl ワークシート
        image_data: PNG/JPEGバイトデータ
        row: 行番号 (1-based)
        col: 列番号 (1-based)
        width_px: 表示幅 (px)
        height_px: 表示高さ (px)
    """
    img = XlImage(io.BytesIO(image_data))
    img.width = width_px
    img.height = height_px

    # セルサイズを取得し、中央配置のオフセットを計算
    cell_w_emu, cell_h_emu = _get_cell_size_emu(ws, row, col)
    img_w_emu = pixels_to_EMU(width_px)
    img_h_emu = pixels_to_EMU(height_px)
    col_off = max(0, (cell_w_emu - img_w_emu) // 2)
    row_off = max(0, (cell_h_emu - img_h_emu) // 2)

    # OneCellAnchorで位置指定 (0-based座標)
    marker = AnchorMarker(
        col=col - 1,
        colOff=col_off,
        row=row - 1,
        rowOff=row_off,
    )

    from openpyxl.drawing.xdr import XDRPositiveSize2D
    anchor = OneCellAnchor(
        _from=marker,
        ext=XDRPositiveSize2D(
            cx=pixels_to_EMU(width_px),
            cy=pixels_to_EMU(height_px),
        ),
    )
    img.anchor = anchor
    ws.add_image(img)
