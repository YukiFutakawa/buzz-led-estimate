"""見積テンプレートExcelへのデータ書き込み

openpyxlはセル値の追跡用にのみ使用し、最終出力はテンプレートZIPに対して
XMLレベルのセル値パッチ + ZIPレベルの画像挿入で生成する。
これによりopenpyxlのload→saveが壊す書式・画像・数式等の問題を回避する。
"""

from __future__ import annotations

import io
import logging
import posixpath
import re
import shutil
import xml.etree.ElementTree as ET
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.utils import get_column_letter

from models import (
    ExistingFixture,
    FloorQuantities,
    LEDProduct,
    MatchResult,
    PropertyInfo,
    QuotationJob,
    SurveyData,
    TEMPLATE_SHEET_MAP,
    INPUT_SHEET,
    SELECTION_SHEET,
    BREAKDOWN_SHEET,
    EXCLUSION_SHEET,
)
from image_handler import (
    LineupImageIndex,
    resize_for_cell,
    prepare_fixture_photo,
    BREAKDOWN_PHOTO_W, BREAKDOWN_PHOTO_H,
    EXCLUSION_PHOTO_W, EXCLUSION_PHOTO_H,
    SELECTION_PHOTO1_W, SELECTION_PHOTO2_W, SELECTION_PHOTO_H,
)

logger = logging.getLogger(__name__)

# ☆入力シートの行ラベル（A〜AD）と対応する行番号
ROW_LABELS = [
    "A", "B", "C", "D", "E", "F", "G", "H", "I", "J",
    "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T",
    "U", "V", "W", "Z", "Y", "Z", "AA", "AB", "AC", "AD",
]

# XML 名前空間定数
NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_PKG = "http://schemas.openxmlformats.org/package/2006/relationships"

# EMU変換定数 (1ピクセル = 9525 EMU at 96 DPI)
EMU_PER_PIXEL = 9525


@dataclass
class ImagePlacement:
    """ZIPレベル画像挿入のための配置情報

    Attributes:
        sheet_name: シート名（例: "選定"）
        image_data: JPEG/PNGバイトデータ
        row: セル行番号 (0-based)
        col: セル列番号 (0-based)
        width_px: 表示幅 (px)
        height_px: 表示高さ (px)
        col_offset_emu: 列オフセット (EMU単位, 中央配置用)
        row_offset_emu: 行オフセット (EMU単位, 中央配置用)
    """
    sheet_name: str
    image_data: bytes
    row: int           # 0-based
    col: int           # 0-based
    width_px: int
    height_px: int
    col_offset_emu: int = 0
    row_offset_emu: int = 0


def _detect_image_format(data: bytes) -> str:
    """画像データのフォーマットを判定"""
    if data[:8] == b'\x89PNG\r\n\x1a\n':
        return "png"
    if data[:2] == b'\xff\xd8':
        return "jpeg"
    # デフォルトはJPEG（resize_for_cellがJPEGを返すため）
    return "jpeg"


def _get_sheet_names(template_name: str) -> dict[str, str]:
    """テンプレート名からシート名マッピングを取得"""
    if template_name in TEMPLATE_SHEET_MAP:
        return TEMPLATE_SHEET_MAP[template_name]
    return TEMPLATE_SHEET_MAP["default"]


def _col_to_idx(col_letter: str) -> int:
    """列文字をopenpyxlの列番号（1始まり）に変換"""
    result = 0
    for c in col_letter.upper():
        result = result * 26 + (ord(c) - ord('A') + 1)
    return result


def _safe_write(
    ws, row: int, col: int, value,
    tracker: dict[str, list[tuple[int, int, object]]] | None = None,
) -> None:
    """結合セルを考慮した安全な書き込み（セル追跡付き）"""
    actual_row, actual_col = row, col

    cell = ws.cell(row=row, column=col)
    if isinstance(cell, openpyxl.cell.cell.MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if (merged_range.min_row <= row <= merged_range.max_row and
                    merged_range.min_col <= col <= merged_range.max_col):
                actual_row = merged_range.min_row
                actual_col = merged_range.min_col
                cell = ws.cell(row=actual_row, column=actual_col)
                break

    cell.value = value

    if tracker is not None:
        sheet_name = ws.title
        if sheet_name not in tracker:
            tracker[sheet_name] = []
        tracker[sheet_name].append((actual_row, actual_col, value))


# ─────────────────────────────────────────────────────────
# テンプレートベース ZIP 再構築
# ─────────────────────────────────────────────────────────

def _get_sheet_xml_map(zf: zipfile.ZipFile) -> dict[str, str]:
    """シート名 → ZIP内XMLパスのマッピングを返す"""
    wb_xml = ET.fromstring(zf.read("xl/workbook.xml"))

    sheet_rids: dict[str, str] = {}
    for elem in wb_xml.iter(f"{{{NS_MAIN}}}sheet"):
        name = elem.get("name")
        rid = elem.get(f"{{{NS_R}}}id")
        if name and rid:
            sheet_rids[name] = rid

    rels_xml = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    rid_targets: dict[str, str] = {}
    for tag in (f"{{{NS_PKG}}}Relationship", "Relationship"):
        for elem in rels_xml.iter(tag):
            rid = elem.get("Id")
            if rid and rid not in rid_targets:
                rid_targets[rid] = elem.get("Target")

    result: dict[str, str] = {}
    for name, rid in sheet_rids.items():
        target = rid_targets.get(rid)
        if target:
            if not target.startswith("/"):
                target = f"xl/{target}"
            else:
                target = target.lstrip("/")
            result[name] = target
    return result


def _get_sheet_drawing_map(zf: zipfile.ZipFile) -> dict[str, str]:
    """シートXMLパス → drawing XMLパスのマッピングを返す

    各シートの.relsファイルを読み、drawing参照を探す。
    例: xl/worksheets/sheet2.xml → xl/drawings/drawing2.xml
    """
    result: dict[str, str] = {}
    sheet_xml_map = _get_sheet_xml_map(zf)

    for sheet_name, sheet_xml_path in sheet_xml_map.items():
        dir_part, file_part = posixpath.split(sheet_xml_path)
        rels_path = posixpath.join(dir_part, "_rels", file_part + ".rels")

        if rels_path not in zf.namelist():
            continue

        rels_text = zf.read(rels_path).decode("utf-8")
        # drawing参照を探す
        for m in re.finditer(
            r'<Relationship[^>]*?'
            r'Type="[^"]*?/drawing"[^>]*?'
            r'Target="([^"]+)"[^>]*?/>',
            rels_text,
        ):
            target = m.group(1)
            # 相対パスを絶対パスに変換
            if not target.startswith("/"):
                drawing_path = posixpath.normpath(
                    posixpath.join(dir_part, target)
                )
            else:
                drawing_path = target.lstrip("/")
            result[sheet_name] = drawing_path
            break  # 1シートに1drawingのみ

    return result


def _xml_escape(text: str) -> str:
    """XML用エスケープ"""
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&apos;")
    )


def _patch_sheet_xml(
    template_xml: bytes,
    writes: list[tuple[int, int, object]],
) -> bytes:
    """テンプレートシートXMLにセル値をパッチして返す

    テンプレートのXMLを文字列レベルで操作し、指定セルの値のみ差し替える。
    スタイル(s属性)は元のまま保持されるため、書式が壊れない。
    文字列値はインライン文字列(t="inlineStr")で書き込み、
    sharedStrings.xmlへの依存を避ける。
    新規セル挿入時は列順序を維持する（列順序違反は修復ダイアログの原因）。
    """
    text = template_xml.decode("utf-8")

    for row_num, col_num, value in writes:
        if value is None:
            continue

        col_letter = get_column_letter(col_num)
        cell_ref = f"{col_letter}{row_num}"

        # 値のXML表現を構築
        if isinstance(value, bool):
            val_xml = "<v>1</v>" if value else "<v>0</v>"
            t_attr = ' t="b"'
        elif isinstance(value, (int, float)):
            val_xml = f"<v>{value}</v>"
            t_attr = ""
        else:
            escaped = _xml_escape(str(value))
            val_xml = f"<is><t>{escaped}</t></is>"
            t_attr = ' t="inlineStr"'

        # 既存セル要素を検索: <c r="C5" ...>...</c> or <c r="C5" .../>
        pattern = re.compile(
            rf'<c\s[^>]*?r="{re.escape(cell_ref)}"[^/]*?/>'
            rf'|<c\s[^>]*?r="{re.escape(cell_ref)}"[^>]*?>.*?</c>',
            re.DOTALL,
        )
        m = pattern.search(text)

        if m:
            # 既存セルのスタイル属性を抽出して保持
            old_cell = m.group(0)
            s_match = re.search(r's="(\d+)"', old_cell)
            s_attr = f' s="{s_match.group(1)}"' if s_match else ""

            new_cell = f'<c r="{cell_ref}"{s_attr}{t_attr}>{val_xml}</c>'
            text = text[: m.start()] + new_cell + text[m.end():]
        else:
            # セルが存在しない → 対応する行に追加（列順序を維持）
            new_c = f'<c r="{cell_ref}"{t_attr}>{val_xml}</c>'

            row_open_pat = re.compile(
                rf'<row\s[^>]*?r="{row_num}"[^>]*?(/?)>'
            )
            rm = row_open_pat.search(text)
            if rm:
                if rm.group(1) == "/":
                    # 自己閉じ行 → 開いてセルを挿入
                    replacement = rm.group(0)[:-2] + ">" + new_c + "</row>"
                    text = text[: rm.start()] + replacement + text[rm.end():]
                else:
                    # 既存の行内で正しい列位置を探す
                    row_tag_end = rm.end()
                    close_pos = text.find("</row>", row_tag_end)
                    if close_pos < 0:
                        close_pos = row_tag_end
                    row_body = text[row_tag_end:close_pos]

                    # 既存セルの列位置をスキャンして挿入位置を決定
                    insert_offset = len(row_body)  # デフォルト: 末尾
                    cell_pat = re.compile(
                        r'<c\s[^>]*?r="([A-Z]+)\d+"'
                    )
                    for cm in cell_pat.finditer(row_body):
                        existing_col = cm.group(1)
                        if _col_to_idx(existing_col) > col_num:
                            insert_offset = cm.start()
                            break

                    abs_insert = row_tag_end + insert_offset
                    text = text[:abs_insert] + new_c + text[abs_insert:]
            else:
                # 行も存在しない → </sheetData> の直前に追加
                sd_end = text.find("</sheetData>")
                if sd_end > 0:
                    new_row = (
                        f'<row r="{row_num}">'
                        f'<c r="{cell_ref}"{t_attr}>{val_xml}</c>'
                        f"</row>"
                    )
                    text = text[:sd_end] + new_row + text[sd_end:]

    return text.encode("utf-8")


def _set_full_calc_on_load(wb_xml_bytes: bytes) -> bytes:
    """workbook.xmlの<calcPr>にfullCalcOnLoad="1"を設定"""
    text = wb_xml_bytes.decode("utf-8")

    calcpr_pattern = re.compile(r"<calcPr\s([^/]*?)/>|<calcPr\s([^>]*?)>")
    m = calcpr_pattern.search(text)

    if m:
        old_tag = m.group(0)
        if "fullCalcOnLoad" in old_tag:
            new_tag = re.sub(
                r'fullCalcOnLoad="[^"]*"',
                'fullCalcOnLoad="1"',
                old_tag,
            )
        else:
            new_tag = old_tag.replace("<calcPr ", '<calcPr fullCalcOnLoad="1" ')
        text = text.replace(old_tag, new_tag)
    else:
        insert_pos = text.rfind("</workbook>")
        if insert_pos > 0:
            text = (
                text[:insert_pos]
                + '<calcPr fullCalcOnLoad="1"/>'
                + text[insert_pos:]
            )

    return text.encode("utf-8")


def _inject_images_into_zip(
    nzf: zipfile.ZipFile,
    tzf: zipfile.ZipFile,
    image_placements: list[ImagePlacement],
    sheet_drawing_map: dict[str, str],
    sheet_xml_map: dict[str, str],
) -> tuple[set[str], dict[str, str], list[str]]:
    """ZIPレベルで画像をdrawing XMLに挿入する

    drawingが存在しないシートには新規drawing XMLを作成し、
    シートrelsとシートXMLにdrawing参照を追加する。

    Returns:
        (modified_paths, sheet_rels_patches, new_content_type_overrides)
    """
    if not image_placements:
        return set(), {}, []

    drawing_images: dict[str, list[tuple[ImagePlacement, int]]] = {}
    media_files: dict[str, bytes] = {}

    existing_media = [
        n for n in tzf.namelist() if n.startswith("xl/media/")
    ]
    media_counter = len(existing_media) + 1

    # 既存drawingの最大番号を取得
    drawing_nums = []
    for n in tzf.namelist():
        dm = re.search(r'drawing(\d+)\.xml$', n)
        if dm:
            drawing_nums.append(int(dm.group(1)))
    next_drawing_num = max(drawing_nums, default=0) + 1

    # drawingが無いシートに新規drawingを作成
    sheets_needing_drawing: dict[str, str] = {}
    sheet_rels_patches: dict[str, str] = {}
    new_content_type_overrides: list[str] = []

    for placement in image_placements:
        drawing_path = sheet_drawing_map.get(placement.sheet_name)
        if not drawing_path:
            if placement.sheet_name in sheets_needing_drawing:
                drawing_path = sheets_needing_drawing[placement.sheet_name]
            else:
                new_drawing_path = f"xl/drawings/drawing{next_drawing_num}.xml"
                sheets_needing_drawing[placement.sheet_name] = new_drawing_path
                sheet_drawing_map[placement.sheet_name] = new_drawing_path
                drawing_path = new_drawing_path

                sheet_xml_path = sheet_xml_map.get(placement.sheet_name)
                if sheet_xml_path:
                    dir_part, file_part = posixpath.split(sheet_xml_path)
                    sheet_rels_path = posixpath.join(
                        dir_part, "_rels", file_part + ".rels"
                    )
                    _existing_rids = []
                    if sheet_rels_path in tzf.namelist():
                        _sr = tzf.read(sheet_rels_path).decode("utf-8")
                        _existing_rids = [
                            int(x) for x in re.findall(
                                r'Id="rId(\d+)"', _sr
                            )
                        ]
                    _rid = max(_existing_rids, default=0) + 1
                    _rid_str = f"rId{_rid}"
                    drawing_rel_target = (
                        f"../drawings/drawing{next_drawing_num}.xml"
                    )
                    sheet_rels_patches[sheet_rels_path] = (
                        f'<Relationship Id="{_rid_str}" '
                        f'Type="http://schemas.openxmlformats.org/officeDocument'
                        f'/2006/relationships/drawing" '
                        f'Target="{drawing_rel_target}"/>'
                    )

                new_content_type_overrides.append(
                    f'<Override PartName="/{new_drawing_path}" '
                    f'ContentType="application/vnd.openxmlformats-'
                    f'officedocument.drawing+xml"/>'
                )
                logger.info(
                    f"新規drawing作成予定: {new_drawing_path} "
                    f"(シート={placement.sheet_name})"
                )
                next_drawing_num += 1

        fmt = _detect_image_format(placement.image_data)
        ext = "png" if fmt == "png" else "jpeg"
        media_name = f"xl/media/image{media_counter}.{ext}"
        media_files[media_name] = placement.image_data

        if drawing_path not in drawing_images:
            drawing_images[drawing_path] = []
        drawing_images[drawing_path].append((placement, media_counter))
        media_counter += 1

    modified_drawings: set[str] = set()

    for drawing_path, placements_with_ids in drawing_images.items():
        # drawingのrelsパスを構築
        dir_part, file_part = posixpath.split(drawing_path)
        rels_path = posixpath.join(dir_part, "_rels", file_part + ".rels")

        # テンプレートのdrawing XMLを読む（新規は空テンプレート）
        if drawing_path in tzf.namelist():
            drawing_xml = tzf.read(drawing_path).decode("utf-8")
        elif drawing_path in sheets_needing_drawing.values():
            drawing_xml = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<xdr:wsDr xmlns:xdr='
                '"http://schemas.openxmlformats.org/drawingml/2006/'
                'spreadsheetDrawing" '
                'xmlns:a='
                '"http://schemas.openxmlformats.org/drawingml/2006/main">'
                '</xdr:wsDr>'
            )
        else:
            logger.warning(f"drawing XMLが見つかりません: {drawing_path}")
            continue

        # テンプレートのdrawing relsを読む（無い場合は空を作成）
        if rels_path in tzf.namelist():
            rels_xml = tzf.read(rels_path).decode("utf-8")
        else:
            rels_xml = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Relationships xmlns='
                '"http://schemas.openxmlformats.org/package/2006/relationships">'
                '</Relationships>'
            )

        # 既存のrId番号の最大値を取得
        existing_rids = [
            int(x) for x in re.findall(r'Id="rId(\d+)"', rels_xml)
        ]
        next_rid = max(existing_rids, default=0) + 1

        # 各画像のoneCellAnchorとrels Relationshipを生成
        new_anchors = []
        new_rels = []

        for placement, media_id in placements_with_ids:
            rid = f"rId{next_rid}"
            next_rid += 1

            fmt = _detect_image_format(placement.image_data)
            ext = "png" if fmt == "png" else "jpeg"

            # drawingからmediaへの相対パス
            media_target = f"../media/image{media_id}.{ext}"

            # Relationship追加
            new_rels.append(
                f'<Relationship Id="{rid}" '
                f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" '
                f'Target="{media_target}"/>'
            )

            # oneCellAnchor XML
            cx = placement.width_px * EMU_PER_PIXEL
            cy = placement.height_px * EMU_PER_PIXEL
            col_off = placement.col_offset_emu
            row_off = placement.row_offset_emu

            anchor_xml = (
                '<xdr:oneCellAnchor>'
                '<xdr:from>'
                f'<xdr:col>{placement.col}</xdr:col>'
                f'<xdr:colOff>{col_off}</xdr:colOff>'
                f'<xdr:row>{placement.row}</xdr:row>'
                f'<xdr:rowOff>{row_off}</xdr:rowOff>'
                '</xdr:from>'
                f'<xdr:ext cx="{cx}" cy="{cy}"/>'
                '<xdr:pic>'
                '<xdr:nvPicPr>'
                f'<xdr:cNvPr id="{media_id + 100}" name="Picture {media_id}"/>'
                '<xdr:cNvPicPr><a:picLocks noChangeAspect="1"/></xdr:cNvPicPr>'
                '</xdr:nvPicPr>'
                '<xdr:blipFill>'
                f'<a:blip xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" r:embed="{rid}"/>'
                '<a:stretch><a:fillRect/></a:stretch>'
                '</xdr:blipFill>'
                '<xdr:spPr>'
                f'<a:xfrm><a:off x="0" y="0"/><a:ext cx="{cx}" cy="{cy}"/></a:xfrm>'
                '<a:prstGeom prst="rect"><a:avLst/></a:prstGeom>'
                '</xdr:spPr>'
                '</xdr:pic>'
                '<xdr:clientData/>'
                '</xdr:oneCellAnchor>'
            )
            new_anchors.append(anchor_xml)

        # drawing XMLに名前空間宣言を確認・追加
        if 'xmlns:a=' not in drawing_xml:
            drawing_xml = drawing_xml.replace(
                'xmlns:xdr=',
                'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" xmlns:xdr=',
            )

        # </xdr:wsDr> の直前にanchorを挿入
        close_tag = "</xdr:wsDr>"
        close_pos = drawing_xml.rfind(close_tag)
        if close_pos >= 0:
            drawing_xml = (
                drawing_xml[:close_pos]
                + "".join(new_anchors)
                + drawing_xml[close_pos:]
            )

        # relsに新しいRelationshipを挿入
        rels_close = "</Relationships>"
        rels_close_pos = rels_xml.rfind(rels_close)
        if rels_close_pos >= 0:
            rels_xml = (
                rels_xml[:rels_close_pos]
                + "".join(new_rels)
                + rels_xml[rels_close_pos:]
            )

        # 変更後のdrawing XMLとrelsをZIPに書き込む
        nzf.writestr(drawing_path, drawing_xml.encode("utf-8"))
        nzf.writestr(rels_path, rels_xml.encode("utf-8"))
        modified_drawings.add(drawing_path)
        modified_drawings.add(rels_path)

        logger.info(
            f"drawing画像挿入: {drawing_path} に "
            f"{len(placements_with_ids)}枚追加"
        )

    # mediaファイルをZIPに追加
    for media_name, data in media_files.items():
        nzf.writestr(media_name, data)

    return modified_drawings, sheet_rels_patches, new_content_type_overrides


def _ensure_content_types_for_images(ct_text: str) -> str:
    """Content_Types.xmlにjpeg/png拡張子の型定義があるか確認し、無ければ追加"""
    needs = []
    if 'Extension="jpeg"' not in ct_text and 'Extension="jpg"' not in ct_text:
        needs.append('<Default Extension="jpeg" ContentType="image/jpeg"/>')
    if 'Extension="png"' not in ct_text:
        needs.append('<Default Extension="png" ContentType="image/png"/>')

    if not needs:
        return ct_text

    # <Types ...> の直後に追加
    insert_pos = ct_text.find(">", ct_text.find("<Types")) + 1
    if insert_pos > 0:
        ct_text = ct_text[:insert_pos] + "".join(needs) + ct_text[insert_pos:]

    return ct_text


def _patch_sheet_rels(rels_text: str, new_rel_xml: str) -> str:
    """シートrelsに新しいRelationshipを追加（rIdは計算済み）"""
    close_tag = "</Relationships>"
    close_pos = rels_text.rfind(close_tag)
    if close_pos >= 0:
        rels_text = (
            rels_text[:close_pos] + new_rel_xml + rels_text[close_pos:]
        )
    return rels_text


def _add_drawing_ref_to_sheet(
    sheet_data: bytes,
    sheet_xml_path: str,
    sheet_rels_patches: dict[str, str],
) -> bytes:
    """シートXMLに<drawing r:id="..."/>要素を追加（まだ無い場合）

    OOXMLスキーマでは<worksheet>直下の要素順序が厳密に定められている。
    <drawing>はこの位置に入る:
      ... → pageSetup → headerFooter → drawing → ... →
      colBreaks → rowBreaks → ... → tableParts → extLst
    つまり<drawing>は pageSetup/headerFooter の後、
    colBreaks/rowBreaks/tableParts/extLst の前に配置する。
    """
    dir_part, file_part = posixpath.split(sheet_xml_path)
    rels_path = posixpath.join(dir_part, "_rels", file_part + ".rels")

    if rels_path not in sheet_rels_patches:
        return sheet_data

    text = sheet_data.decode("utf-8") if isinstance(sheet_data, bytes) else sheet_data

    if re.search(r'<drawing\s', text):
        return text.encode("utf-8") if isinstance(sheet_data, bytes) else sheet_data

    patch_xml = sheet_rels_patches[rels_path]
    rid_match = re.search(r'Id="(rId\d+)"', patch_xml)
    rid_val = rid_match.group(1) if rid_match else "rId1"

    drawing_elem = f'<drawing r:id="{rid_val}"/>'

    # OOXMLスキーマ順序 (CT_Worksheet):
    # ... → pageSetup → headerFooter → rowBreaks → colBreaks →
    # customProperties → cellWatches → ignoredErrors → smartTags →
    # drawing → legacyDrawing → legacyDrawingHF → drawingHF →
    # picture → oleObjects → controls → webPublishItems →
    # tableParts → extLst → </worksheet>
    #
    # <drawing>より「後」に来る要素の直前に挿入する
    _AFTER_DRAWING = [
        r'<legacyDrawing[\s/>]',
        r'<legacyDrawingHF[\s/>]',
        r'<drawingHF[\s/>]',
        r'<picture[\s/>]',
        r'<oleObjects[\s/>]',
        r'<controls[\s/>]',
        r'<webPublishItems[\s/>]',
        r'<tableParts[\s/>]',
        r'<extLst[\s/>]',
        r'</worksheet>',
    ]

    insert_pos = None
    for pattern in _AFTER_DRAWING:
        m = re.search(pattern, text)
        if m:
            insert_pos = m.start()
            break

    if insert_pos is not None:
        text = text[:insert_pos] + drawing_elem + text[insert_pos:]
    else:
        # フォールバック: </worksheet>の前
        ws_close_pos = text.rfind("</worksheet>")
        if ws_close_pos >= 0:
            text = text[:ws_close_pos] + drawing_elem + text[ws_close_pos:]

    return text.encode("utf-8") if isinstance(sheet_data, bytes) else text


def _rebuild_from_template(
    template_path: Path,
    output_path: Path,
    modified_sheet_names: set[str],
    cell_writes: dict[str, list[tuple[int, int, object]]],
    image_placements: list[ImagePlacement] | None = None,
) -> None:
    """テンプレートZIPベースで出力ファイルを完全再構築

    openpyxlの出力を完全に破棄し、テンプレートZIPに対して:
    1. セル値パッチ（XMLレベル）
    2. 画像挿入（ZIPレベル: drawing XML + rels + media）
    3. calcChain除去 + fullCalcOnLoad設定
    のみ行う。openpyxlの保存結果は一切使わない。
    """
    with zipfile.ZipFile(str(template_path), "r") as tzf:
        tpl_sheet_map = _get_sheet_xml_map(tzf)
        sheet_drawing_map = _get_sheet_drawing_map(tzf)

        # 変更シートXMLにセル値をパッチ
        patched_sheets: dict[str, bytes] = {}
        for sheet_name in modified_sheet_names:
            xml_path = tpl_sheet_map.get(sheet_name)
            if not xml_path:
                continue
            tpl_xml = tzf.read(xml_path)
            writes = cell_writes.get(sheet_name, [])
            if writes:
                patched = _patch_sheet_xml(tpl_xml, writes)
                logger.info(
                    f"シートXMLパッチ: {sheet_name} "
                    f"({len(writes)}セル書き込み)"
                )
            else:
                patched = tpl_xml
            patched_sheets[xml_path] = patched

        # スキップ対象ファイル
        SKIP_FILES = {"xl/calcChain.xml"}

        # 新しいZIPを構築
        temp_path = output_path.with_suffix(".tmp.xlsx")
        with zipfile.ZipFile(
            str(temp_path), "w", zipfile.ZIP_DEFLATED,
        ) as nzf:
            # 画像をdrawing XMLに挿入
            modified_paths: set[str] = set()
            sheet_rels_patches: dict[str, str] = {}
            new_ct_overrides: list[str] = []
            if image_placements:
                modified_paths, sheet_rels_patches, new_ct_overrides = (
                    _inject_images_into_zip(
                        nzf, tzf, image_placements,
                        sheet_drawing_map, tpl_sheet_map,
                    )
                )

            written: set[str] = set(modified_paths)

            for item in tzf.infolist():
                fname = item.filename

                if fname in SKIP_FILES:
                    logger.debug(f"スキップ: {fname}")
                    written.add(fname)
                    continue

                if fname in modified_paths:
                    continue

                if fname in patched_sheets:
                    sheet_data = patched_sheets[fname]
                    sheet_data = _add_drawing_ref_to_sheet(
                        sheet_data, fname, sheet_rels_patches,
                    )
                    nzf.writestr(item, sheet_data)
                elif fname in sheet_rels_patches:
                    rels_data = tzf.read(fname).decode("utf-8")
                    rels_data = _patch_sheet_rels(
                        rels_data, sheet_rels_patches[fname],
                    )
                    nzf.writestr(item, rels_data.encode("utf-8"))
                elif fname == "xl/workbook.xml":
                    wb_data = _set_full_calc_on_load(tzf.read(fname))
                    nzf.writestr(item, wb_data)
                elif fname == "[Content_Types].xml":
                    ct_text = tzf.read(fname).decode("utf-8")
                    ct_text = re.sub(
                        r'<Override[^>]*?PartName="[^"]*calcChain[^"]*"[^>]*/?>',
                        "", ct_text,
                    )
                    if image_placements:
                        ct_text = _ensure_content_types_for_images(ct_text)
                    if new_ct_overrides:
                        ct_close_pos = ct_text.rfind("</Types>")
                        if ct_close_pos >= 0:
                            ct_text = (
                                ct_text[:ct_close_pos]
                                + "".join(new_ct_overrides)
                                + ct_text[ct_close_pos:]
                            )
                    nzf.writestr(item, ct_text.encode("utf-8"))
                else:
                    # 未パッチシートでもdrawing参照が必要な場合
                    needs_drawing = False
                    for sn, sp in tpl_sheet_map.items():
                        if sp == fname:
                            dir_p, file_p = posixpath.split(fname)
                            rels_p = posixpath.join(
                                dir_p, "_rels", file_p + ".rels"
                            )
                            if rels_p in sheet_rels_patches:
                                needs_drawing = True
                            break
                    if needs_drawing:
                        sdata = _add_drawing_ref_to_sheet(
                            tzf.read(fname), fname,
                            sheet_rels_patches,
                        )
                        nzf.writestr(item, sdata)
                    else:
                        nzf.writestr(item, tzf.read(fname))

                written.add(fname)

            # 新規シートrels（テンプレートに無い場合）
            for rels_path, rel_xml in sheet_rels_patches.items():
                if rels_path not in written:
                    new_rels = (
                        '<?xml version="1.0" encoding="UTF-8" '
                        'standalone="yes"?>'
                        '<Relationships xmlns='
                        '"http://schemas.openxmlformats.org/package/'
                        '2006/relationships">'
                        f'{rel_xml}'
                        '</Relationships>'
                    )
                    nzf.writestr(rels_path, new_rels.encode("utf-8"))
                    written.add(rels_path)

    temp_path.replace(output_path)
    logger.info(
        f"テンプレートベース再構築完了: "
        f"{len(patched_sheets)}シートパッチ, "
        f"{len(image_placements or [])}画像挿入"
    )


# ─────────────────────────────────────────────────────────
# メインクラス
# ─────────────────────────────────────────────────────────

class ExcelWriter:
    """見積テンプレートへの書き込みエンジン"""

    def __init__(
        self,
        template_dir: Path,
        image_index: Optional[LineupImageIndex] = None,
    ):
        self.template_dir = template_dir
        self.image_index = image_index

    def write_quotation(self, job: QuotationJob) -> Path:
        """QuotationJobの内容をテンプレートに書き込み、出力パスを返す"""

        # テンプレートファイルを特定
        template_path = self._find_template(job.template_name)
        if template_path is None:
            raise FileNotFoundError(
                f"テンプレートが見つかりません: {job.template_name}"
            )

        # 出力先にテンプレートをコピー
        if job.output_path is None:
            output_dir = self.template_dir.parent / "output"
            output_dir.mkdir(exist_ok=True)
            safe_name = job.survey.property_info.name or "unnamed"
            job.output_path = output_dir / f"【LED導入ｼﾐｭﾚｰｼｮﾝ】{safe_name}.xlsx"

        shutil.copy2(template_path, job.output_path)

        # セル書き込み追跡用辞書
        cell_tracker: dict[str, list[tuple[int, int, object]]] = {}

        # openpyxlでセル書き込み追跡（結合セル検出のため）
        wb = openpyxl.load_workbook(job.output_path)
        sheet_names = _get_sheet_names(job.template_name)

        # ☆入力シートへの書き込み
        ws_input = wb[sheet_names["input"]]
        self._write_property_info(ws_input, job.survey.property_info,
                                  cell_tracker)
        self._write_fixture_rows(ws_input, job.matches, cell_tracker)
        self._write_excluded_rows(ws_input, job.survey.excluded_fixtures,
                                  cell_tracker)

        # 選定シートへの書き込み
        ws_selection = wb[sheet_names["selection"]]
        self._write_selection_sheet(ws_selection, job.matches, cell_tracker)

        # 書き込み対象シート名を記録
        modified_sheets = {
            sheet_names["input"],
            sheet_names["selection"],
        }

        wb.close()
        # openpyxlの保存結果は使わない（_rebuild_from_templateで破棄される）

        # 画像をImagePlacementとして収集
        image_placements: list[ImagePlacement] = []
        if self.image_index:
            self._collect_selection_photos(
                sheet_names["selection"], job.matches, image_placements,
            )

            breakdown_name = sheet_names.get("breakdown")
            if breakdown_name:
                self._collect_breakdown_photos(
                    breakdown_name, job.matches, job.survey,
                    image_placements,
                )

            exclusion_name = sheet_names.get("exclusion")
            if exclusion_name:
                self._collect_exclusion_photos(
                    exclusion_name, job.survey.excluded_fixtures,
                    image_placements,
                )

        # テンプレートZIPベースで出力ファイルを完全再構築
        _rebuild_from_template(
            template_path, job.output_path,
            modified_sheets, cell_tracker,
            image_placements,
        )

        logger.info(f"見積ファイル出力完了: {job.output_path}")
        return job.output_path

    def _find_template(self, template_name: str) -> Optional[Path]:
        """テンプレート名からファイルパスを検索"""
        for f in self.template_dir.iterdir():
            if f.suffix == ".xlsx" and template_name in f.name:
                return f
        return None

    def _write_property_info(
        self, ws, info: PropertyInfo,
        tracker: dict[str, list[tuple[int, int, object]]],
    ) -> None:
        """物件情報をヘッダーエリアに書き込み"""
        if info.name:
            _safe_write(ws, 5, _col_to_idx("C"), info.name, tracker)
        if info.address:
            _safe_write(ws, 6, _col_to_idx("C"), info.address, tracker)
        if info.unlock_code:
            _safe_write(ws, 7, _col_to_idx("C"), info.unlock_code, tracker)
        if info.distribution_board:
            _safe_write(ws, 8, _col_to_idx("C"), info.distribution_board,
                        tracker)
        if info.special_notes:
            _safe_write(ws, 9, _col_to_idx("C"), info.special_notes, tracker)

        logger.info(f"物件情報書き込み: {info.name} / {info.address}")

    def _write_fixture_rows(
        self, ws, matches: list[MatchResult],
        tracker: dict[str, list[tuple[int, int, object]]],
    ) -> None:
        """☆入力シートのデータ行（Row 16-45）に器具データを書き込み"""
        start_row = INPUT_SHEET["data_start_row"]  # 16

        for i, match in enumerate(matches):
            if i >= 30:
                logger.warning(
                    "器具種別が30を超えました。超過分はスキップします。"
                )
                break

            row = start_row + i
            fixture = match.fixture

            _safe_write(ws, row, _col_to_idx("C"), fixture.location, tracker)
            _safe_write(ws, row, _col_to_idx("D"), fixture.fixture_type,
                        tracker)
            if fixture.survey_notes:
                _safe_write(ws, row, _col_to_idx("E"), fixture.survey_notes,
                            tracker)
            if fixture.construction_notes:
                _safe_write(ws, row, _col_to_idx("F"),
                            fixture.construction_notes, tracker)
            if match.category_key:
                _safe_write(ws, row, _col_to_idx("G"), match.category_key,
                            tracker)
            if fixture.daily_hours > 0:
                _safe_write(ws, row, _col_to_idx("I"), fixture.daily_hours,
                            tracker)
            if fixture.adjusted_power_w > 0:
                _safe_write(ws, row, _col_to_idx("K"),
                            fixture.adjusted_power_w, tracker)

            # L列: =SUM(M:V) 数式あり → 書き込まない

            # M-V列: 各階数量（1F〜10F）
            floor_start_col = _col_to_idx("M")
            for floor_idx, qty in enumerate(
                fixture.quantities.to_list(10)
            ):
                if qty > 0:
                    _safe_write(ws, row, floor_start_col + floor_idx, qty,
                                tracker)

            if match.construction_unit_price > 0:
                _safe_write(ws, row, _col_to_idx("AE"),
                            match.construction_unit_price, tracker)

        logger.info(f"器具データ {len(matches)}行を書き込みました")

    def _write_excluded_rows(
        self, ws, excluded: list[ExistingFixture],
        tracker: dict[str, list[tuple[int, int, object]]],
    ) -> None:
        """☆入力シートのLED済みセクション（Row 49+）に除外データを書き込み"""
        start_row = INPUT_SHEET["excluded_start_row"]  # 49

        for i, fixture in enumerate(excluded):
            if i >= 10:
                break

            row = start_row + i
            _safe_write(ws, row, _col_to_idx("C"), fixture.location, tracker)
            _safe_write(ws, row, _col_to_idx("D"), fixture.fixture_type,
                        tracker)
            if fixture.survey_notes:
                _safe_write(ws, row, _col_to_idx("E"), fixture.survey_notes,
                            tracker)

            # L列: =SUM(M:V) 数式あり → 直接書き込まない
            # M-V列: 各階数量（1F〜10F）を書き込みL列は自動計算
            floor_start_col = _col_to_idx("M")
            for floor_idx, qty in enumerate(
                fixture.quantities.to_list(10)
            ):
                if qty > 0:
                    _safe_write(ws, row, floor_start_col + floor_idx, qty,
                                tracker)

            if fixture.exclusion_reason:
                _safe_write(ws, row, _col_to_idx("W"),
                            fixture.exclusion_reason, tracker)
            if fixture.exclusion_advice:
                _safe_write(ws, row, _col_to_idx("AC"),
                            fixture.exclusion_advice, tracker)

        if excluded:
            logger.info(f"除外データ {len(excluded)}行を書き込みました")

    def _write_selection_sheet(
        self, ws, matches: list[MatchResult],
        tracker: dict[str, list[tuple[int, int, object]]],
    ) -> None:
        """選定シートにLED商品仕様を書き込み"""
        start_row = SELECTION_SHEET["data_start_row"]  # 3

        seen_keys: set[str] = set()
        unique_matches: list[MatchResult] = []
        for match in matches:
            if match.category_key and match.category_key not in seen_keys:
                seen_keys.add(match.category_key)
                unique_matches.append(match)

        for i, match in enumerate(unique_matches):
            if i >= 30:
                break
            if match.led_product is None:
                continue

            row = start_row + i
            led = match.led_product

            _safe_write(ws, row, _col_to_idx("C"), match.category_key,
                        tracker)
            _safe_write(ws, row, _col_to_idx("D"), led.lighting_color,
                        tracker)
            _safe_write(ws, row, _col_to_idx("E"), led.fixture_color,
                        tracker)
            _safe_write(ws, row, _col_to_idx("F"), led.fixture_size,
                        tracker)
            if led.power_w > 0:
                _safe_write(ws, row, _col_to_idx("G"), led.power_w, tracker)
            _safe_write(ws, row, _col_to_idx("H"), led.lumens, tracker)
            if led.list_price_total > 0:
                _safe_write(ws, row, _col_to_idx("I"), led.list_price_total,
                            tracker)
            if led.purchase_price_total > 0:
                _safe_write(ws, row, _col_to_idx("J"),
                            led.purchase_price_total, tracker)
            _safe_write(ws, row, _col_to_idx("K"),
                        "〇" if led.is_waterproof else "✕", tracker)
            _safe_write(ws, row, _col_to_idx("L"), led.bulb_type, tracker)
            _safe_write(ws, row, _col_to_idx("M"), led.manufacturer, tracker)
            _safe_write(ws, row, _col_to_idx("N"), led.watt_equivalent,
                        tracker)
            _safe_write(ws, row, _col_to_idx("O"), led.model_number, tracker)
            if led.model_price > 0:
                _safe_write(ws, row, _col_to_idx("P"), led.model_price,
                            tracker)
            if led.model_purchase > 0:
                _safe_write(ws, row, _col_to_idx("Q"), led.model_purchase,
                            tracker)
            if led.model_number_2:
                _safe_write(ws, row, _col_to_idx("R"), led.model_number_2,
                            tracker)
            if led.model_price_2:
                _safe_write(ws, row, _col_to_idx("S"), led.model_price_2,
                            tracker)
            if led.model_purchase_2:
                _safe_write(ws, row, _col_to_idx("T"), led.model_purchase_2,
                            tracker)
            if led.model_number_3:
                _safe_write(ws, row, _col_to_idx("U"), led.model_number_3,
                            tracker)
            if led.model_price_3:
                _safe_write(ws, row, _col_to_idx("V"), led.model_price_3,
                            tracker)
            if led.model_purchase_3:
                _safe_write(ws, row, _col_to_idx("W"), led.model_purchase_3,
                            tracker)
            if led.model_number_4:
                _safe_write(ws, row, _col_to_idx("X"), led.model_number_4,
                            tracker)
            if led.model_price_4:
                _safe_write(ws, row, _col_to_idx("Y"), led.model_price_4,
                            tracker)
            if led.model_purchase_4:
                _safe_write(ws, row, _col_to_idx("Z"), led.model_purchase_4,
                            tracker)
            if led.power_detail > 0:
                _safe_write(ws, row, _col_to_idx("AA"), led.power_detail,
                            tracker)
            _safe_write(ws, row, _col_to_idx("AB"), led.lumens_detail,
                        tracker)
            _safe_write(ws, row, _col_to_idx("AC"), led.material, tracker)
            _safe_write(ws, row, _col_to_idx("AD"), led.color_options,
                        tracker)
            _safe_write(ws, row, _col_to_idx("AE"),
                        led.lighting_color_options, tracker)
            _safe_write(ws, row, _col_to_idx("AF"), led.lifespan, tracker)
            _safe_write(ws, row, _col_to_idx("AG"), led.replacement_method,
                        tracker)
            _safe_write(ws, row, _col_to_idx("AH"), led.socket, tracker)

        logger.info(
            f"選定データ {len(unique_matches)}カテゴリを書き込みました"
        )

    # ===== 写真収集メソッド（ZIPレベル挿入用） =====

    def _collect_selection_photos(
        self,
        sheet_name: str,
        matches: list[MatchResult],
        placements: list[ImagePlacement],
    ) -> None:
        """選定シートのLED商品写真をImagePlacementとして収集"""
        start_row = SELECTION_SHEET["data_start_row"]  # 3

        seen_keys: set[str] = set()
        unique_matches: list[MatchResult] = []
        for match in matches:
            if match.category_key and match.category_key not in seen_keys:
                seen_keys.add(match.category_key)
                unique_matches.append(match)

        count = 0
        for i, match in enumerate(unique_matches):
            if i >= 30:
                break
            if match.led_product is None:
                continue

            row_0based = start_row + i - 1  # 0-based

            # 写真① (A列, col=0)
            img1_data = self.image_index.get_product_image(
                match.led_product, photo_num=1,
            )
            if img1_data:
                try:
                    resized = resize_for_cell(
                        img1_data, SELECTION_PHOTO1_W, SELECTION_PHOTO_H,
                    )
                    placements.append(ImagePlacement(
                        sheet_name=sheet_name,
                        image_data=resized,
                        row=row_0based,
                        col=0,  # A列
                        width_px=SELECTION_PHOTO1_W,
                        height_px=SELECTION_PHOTO_H,
                    ))
                    count += 1
                except Exception as e:
                    logger.warning(f"選定写真①収集エラー (row={start_row+i}): {e}")

            # 写真② (B列, col=1)
            img2_data = self.image_index.get_product_image(
                match.led_product, photo_num=2,
            )
            if img2_data:
                try:
                    resized = resize_for_cell(
                        img2_data, SELECTION_PHOTO2_W, SELECTION_PHOTO_H,
                    )
                    placements.append(ImagePlacement(
                        sheet_name=sheet_name,
                        image_data=resized,
                        row=row_0based,
                        col=1,  # B列
                        width_px=SELECTION_PHOTO2_W,
                        height_px=SELECTION_PHOTO_H,
                    ))
                    count += 1
                except Exception as e:
                    logger.warning(f"選定写真②収集エラー (row={start_row+i}): {e}")

        logger.info(f"選定シート写真収集: {count}枚")

    def _collect_breakdown_photos(
        self,
        sheet_name: str,
        matches: list[MatchResult],
        survey: SurveyData,
        placements: list[ImagePlacement],
    ) -> None:
        """⑩内訳シートの写真をImagePlacementとして収集"""
        existing_row = BREAKDOWN_SHEET["existing_photo_row"]  # 7
        led_row = BREAKDOWN_SHEET["led_photo_row"]  # 14
        photo_w = BREAKDOWN_PHOTO_W
        photo_h = BREAKDOWN_PHOTO_H

        count = 0
        for i, match in enumerate(matches):
            if i >= 20:
                break

            col_1based = _col_to_idx("B") + i
            col_0based = col_1based - 1

            # 既存器具写真
            fixture = match.fixture
            if fixture.photo_paths:
                photo_path = fixture.photo_paths[0]
                if photo_path.exists():
                    try:
                        resized = prepare_fixture_photo(
                            photo_path, photo_w, photo_h,
                        )
                        placements.append(ImagePlacement(
                            sheet_name=sheet_name,
                            image_data=resized,
                            row=existing_row - 1,  # 0-based
                            col=col_0based,
                            width_px=photo_w,
                            height_px=photo_h,
                        ))
                        count += 1
                    except Exception as e:
                        logger.warning(
                            f"内訳 既存写真収集エラー (col={col_1based}): {e}"
                        )

            # LED商品写真
            if match.led_product:
                img_data = self.image_index.get_product_image(
                    match.led_product, photo_num=1,
                )
                if img_data:
                    try:
                        resized = resize_for_cell(
                            img_data, photo_w, photo_h,
                        )
                        placements.append(ImagePlacement(
                            sheet_name=sheet_name,
                            image_data=resized,
                            row=led_row - 1,  # 0-based
                            col=col_0based,
                            width_px=photo_w,
                            height_px=photo_h,
                        ))
                        count += 1
                    except Exception as e:
                        logger.warning(
                            f"内訳 LED写真収集エラー (col={col_1based}): {e}"
                        )

        logger.info(f"⑩内訳シート写真収集: {count}枚")

    def _collect_exclusion_photos(
        self,
        sheet_name: str,
        excluded: list[ExistingFixture],
        placements: list[ImagePlacement],
    ) -> None:
        """⑪除外シートの写真をImagePlacementとして収集"""
        blocks = EXCLUSION_SHEET["blocks"]
        photo_w = EXCLUSION_PHOTO_W
        photo_h = EXCLUSION_PHOTO_H

        count = 0
        for i, fixture in enumerate(excluded):
            if i >= len(blocks):
                break

            block = blocks[i]
            col_1based = _col_to_idx(block["photo_col"])
            row_1based = block["photo_row"]

            if fixture.photo_paths:
                photo_path = fixture.photo_paths[0]
                if photo_path.exists():
                    try:
                        resized = prepare_fixture_photo(
                            photo_path, photo_w, photo_h,
                        )
                        placements.append(ImagePlacement(
                            sheet_name=sheet_name,
                            image_data=resized,
                            row=row_1based - 1,  # 0-based
                            col=col_1based - 1,  # 0-based
                            width_px=photo_w,
                            height_px=photo_h,
                        ))
                        count += 1
                    except Exception as e:
                        logger.warning(
                            f"除外写真収集エラー (block={i}): {e}"
                        )

        if count > 0:
            logger.info(f"⑪除外シート写真収集: {count}枚")
