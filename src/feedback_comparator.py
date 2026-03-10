# -*- coding: utf-8 -*-
"""AI生成見積り vs 人間修正版の差分比較

AI生成Excelと人間が修正した正解Excelをセルレベルで比較し、
構造化されたフィードバックデータを生成する。

Usage:
    from feedback_comparator import FeedbackComparator
    comparator = FeedbackComparator()
    report = comparator.compare(ai_path, correct_path)
    report.save_json(output_path)
"""

from __future__ import annotations

import json
import logging
import unicodedata
from dataclasses import dataclass, field, asdict
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl

from models import INPUT_SHEET, SELECTION_SHEET

logger = logging.getLogger(__name__)


# ============================================================
# データクラス
# ============================================================

@dataclass
class CellDiff:
    """1セルの差分"""
    sheet: str
    cell: str          # "C5", "G16" etc.
    field_name: str    # 人間可読フィールド名
    ai_value: str
    correct_value: str
    severity: str      # "critical", "major", "minor"

    @property
    def is_empty_vs_empty(self) -> bool:
        return not self.ai_value and not self.correct_value


@dataclass
class FixtureRowDiff:
    """1器具行の差分まとめ"""
    row_number: int
    row_label: str     # A, B, C...
    fixture_type_ai: str
    fixture_type_correct: str
    diffs: list[CellDiff] = field(default_factory=list)
    status: str = "modified"  # "modified", "missing_in_ai", "extra_in_ai", "match"


@dataclass
class SelectionRowDiff:
    """選定シート1行の差分"""
    row_number: int
    product_key_ai: str
    product_key_correct: str
    diffs: list[CellDiff] = field(default_factory=list)
    status: str = "modified"


@dataclass
class FeedbackReport:
    """比較結果レポート"""
    ai_file: str
    correct_file: str
    timestamp: str = ""
    property_name: str = ""

    # ☆入力: ヘッダー差分
    header_diffs: list[CellDiff] = field(default_factory=list)

    # ☆入力: 器具行差分
    fixture_diffs: list[FixtureRowDiff] = field(default_factory=list)

    # ☆入力: 除外行差分
    excluded_diffs: list[FixtureRowDiff] = field(default_factory=list)

    # 選定シート差分
    selection_diffs: list[SelectionRowDiff] = field(default_factory=list)

    # サマリー統計
    total_fixtures_ai: int = 0
    total_fixtures_correct: int = 0
    total_excluded_ai: int = 0
    total_excluded_correct: int = 0
    total_products_ai: int = 0
    total_products_correct: int = 0

    @property
    def total_diffs(self) -> int:
        count = len(self.header_diffs)
        for fd in self.fixture_diffs:
            count += len(fd.diffs)
        for ed in self.excluded_diffs:
            count += len(ed.diffs)
        for sd in self.selection_diffs:
            count += len(sd.diffs)
        return count

    @property
    def fixture_match_rate(self) -> float:
        if not self.fixture_diffs:
            return 0.0
        matched = sum(1 for f in self.fixture_diffs if f.status == "match")
        return matched / len(self.fixture_diffs)

    @property
    def led_selection_match_rate(self) -> float:
        """G列（LED選定）の一致率"""
        total = 0
        matched = 0
        for fd in self.fixture_diffs:
            if fd.status in ("modified", "match"):
                total += 1
                g_diffs = [d for d in fd.diffs if d.field_name == "LED選定(G列)"]
                if not g_diffs:
                    matched += 1
        return matched / total if total > 0 else 0.0

    def save_json(self, path: Path) -> None:
        """JSON形式で保存"""
        data = {
            "ai_file": self.ai_file,
            "correct_file": self.correct_file,
            "timestamp": self.timestamp,
            "property_name": self.property_name,
            "summary": {
                "total_diffs": self.total_diffs,
                "fixture_match_rate": round(self.fixture_match_rate, 3),
                "led_selection_match_rate": round(self.led_selection_match_rate, 3),
                "fixtures": {
                    "ai": self.total_fixtures_ai,
                    "correct": self.total_fixtures_correct,
                },
                "excluded": {
                    "ai": self.total_excluded_ai,
                    "correct": self.total_excluded_correct,
                },
                "products": {
                    "ai": self.total_products_ai,
                    "correct": self.total_products_correct,
                },
            },
            "header_diffs": [asdict(d) for d in self.header_diffs],
            "fixture_diffs": [
                {
                    "row_number": fd.row_number,
                    "row_label": fd.row_label,
                    "fixture_type_ai": fd.fixture_type_ai,
                    "fixture_type_correct": fd.fixture_type_correct,
                    "status": fd.status,
                    "diffs": [asdict(d) for d in fd.diffs],
                }
                for fd in self.fixture_diffs
            ],
            "excluded_diffs": [
                {
                    "row_number": ed.row_number,
                    "row_label": ed.row_label,
                    "fixture_type_ai": ed.fixture_type_ai,
                    "fixture_type_correct": ed.fixture_type_correct,
                    "status": ed.status,
                    "diffs": [asdict(d) for d in ed.diffs],
                }
                for ed in self.excluded_diffs
            ],
            "selection_diffs": [
                {
                    "row_number": sd.row_number,
                    "product_key_ai": sd.product_key_ai,
                    "product_key_correct": sd.product_key_correct,
                    "status": sd.status,
                    "diffs": [asdict(d) for d in sd.diffs],
                }
                for sd in self.selection_diffs
            ],
        }
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(
            json.dumps(data, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        logger.info(f"フィードバック保存: {path}")

    def print_summary(self) -> str:
        """コンソール表示用サマリー"""
        lines = []
        lines.append(f"{'='*60}")
        lines.append(f"  フィードバック比較レポート")
        lines.append(f"{'='*60}")
        lines.append(f"  物件: {self.property_name}")
        lines.append(f"  AI出力: {self.ai_file}")
        lines.append(f"  正解:   {self.correct_file}")
        lines.append(f"  日時:   {self.timestamp}")
        lines.append(f"")
        lines.append(f"  --- 全体統計 ---")
        lines.append(f"  差分セル数: {self.total_diffs}")
        lines.append(f"  器具行一致率: {self.fixture_match_rate:.1%}")
        lines.append(f"  LED選定一致率: {self.led_selection_match_rate:.1%}")
        lines.append(f"  器具数: AI={self.total_fixtures_ai} / "
                     f"正解={self.total_fixtures_correct}")
        lines.append(f"  除外数: AI={self.total_excluded_ai} / "
                     f"正解={self.total_excluded_correct}")

        # ヘッダー差分
        if self.header_diffs:
            lines.append(f"\n  --- ヘッダー差分 ({len(self.header_diffs)}件) ---")
            for d in self.header_diffs:
                lines.append(f"  [{d.severity}] {d.field_name}")
                lines.append(f"    AI:  {d.ai_value[:60]}")
                lines.append(f"    正解: {d.correct_value[:60]}")

        # 器具行差分（重要な差分のみ表示）
        critical_fixture_diffs = [
            fd for fd in self.fixture_diffs if fd.status != "match"
        ]
        if critical_fixture_diffs:
            lines.append(f"\n  --- 器具行差分 ({len(critical_fixture_diffs)}件) ---")
            for fd in critical_fixture_diffs:
                status_mark = {
                    "modified": "[修正]",
                    "missing_in_ai": "[AI欠落]",
                    "extra_in_ai": "[AI余分]",
                }.get(fd.status, fd.status)
                lines.append(f"\n  {status_mark} Row {fd.row_number} ({fd.row_label})")
                ft = fd.fixture_type_correct or fd.fixture_type_ai
                lines.append(f"    器具: {ft}")
                for d in fd.diffs:
                    lines.append(f"    [{d.severity}] {d.field_name}: "
                                f"AI「{d.ai_value}」→ 正解「{d.correct_value}」")

        # 選定差分
        sel_diffs = [sd for sd in self.selection_diffs if sd.status != "match"]
        if sel_diffs:
            lines.append(f"\n  --- 選定シート差分 ({len(sel_diffs)}件) ---")
            for sd in sel_diffs:
                key = sd.product_key_correct or sd.product_key_ai
                lines.append(f"\n  Row {sd.row_number}: {key}")
                for d in sd.diffs:
                    lines.append(f"    [{d.severity}] {d.field_name}: "
                                f"AI「{d.ai_value}」→ 正解「{d.correct_value}」")

        return "\n".join(lines)


# ============================================================
# 比較エンジン
# ============================================================

ROW_LABELS = [
    "A", "B", "C", "D", "E", "F", "G", "H", "I", "J",
    "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T",
    "U", "V", "W", "X", "Y", "Z", "AA", "AB", "AC", "AD",
]


class FeedbackComparator:
    """AI生成見積り vs 人間修正版の比較エンジン"""

    def compare(self, ai_path: Path, correct_path: Path) -> FeedbackReport:
        """2つのExcelファイルを比較"""
        logger.info(f"比較開始: {ai_path.name} vs {correct_path.name}")

        wb_ai = openpyxl.load_workbook(ai_path, data_only=True)
        wb_correct = openpyxl.load_workbook(correct_path, data_only=True)

        ws_ai = self._find_sheet(wb_ai, "入力")
        ws_correct = self._find_sheet(wb_correct, "入力")

        if not ws_ai or not ws_correct:
            raise ValueError("☆入力シートが見つかりません")

        report = FeedbackReport(
            ai_file=ai_path.name,
            correct_file=correct_path.name,
            timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            property_name=self._safe_str(ws_correct["C5"].value),
        )

        # ヘッダー比較
        report.header_diffs = self._compare_header(ws_ai, ws_correct)

        # 器具行比較
        report.fixture_diffs = self._compare_fixtures(ws_ai, ws_correct)
        report.total_fixtures_ai = sum(
            1 for fd in report.fixture_diffs
            if fd.status in ("modified", "match", "extra_in_ai")
        )
        report.total_fixtures_correct = sum(
            1 for fd in report.fixture_diffs
            if fd.status in ("modified", "match", "missing_in_ai")
        )

        # 除外行比較
        report.excluded_diffs = self._compare_excluded(ws_ai, ws_correct)
        report.total_excluded_ai = sum(
            1 for ed in report.excluded_diffs
            if ed.status in ("modified", "match", "extra_in_ai")
        )
        report.total_excluded_correct = sum(
            1 for ed in report.excluded_diffs
            if ed.status in ("modified", "match", "missing_in_ai")
        )

        # 選定シート比較
        ws_sel_ai = self._find_sheet(wb_ai, "選定")
        ws_sel_correct = self._find_sheet(wb_correct, "選定")
        if ws_sel_ai and ws_sel_correct:
            report.selection_diffs = self._compare_selection(
                ws_sel_ai, ws_sel_correct)
            report.total_products_ai = sum(
                1 for sd in report.selection_diffs
                if sd.status in ("modified", "match", "extra_in_ai")
            )
            report.total_products_correct = sum(
                1 for sd in report.selection_diffs
                if sd.status in ("modified", "match", "missing_in_ai")
            )

        wb_ai.close()
        wb_correct.close()

        logger.info(f"比較完了: 差分{report.total_diffs}件")
        return report

    def compare_folder(
        self,
        ai_dir: Path,
        correct_dir: Path,
    ) -> list[FeedbackReport]:
        """フォルダ内の対応するファイルを一括比較

        ファイル名のマッチングロジック:
        - 正解ファイル名の物件名部分がAI出力ファイル名に含まれるかチェック
        """
        reports = []
        correct_files = {
            f.stem: f
            for f in correct_dir.glob("*.xlsx")
            if not f.name.startswith("~$")
        }

        for ai_file in sorted(ai_dir.glob("*.xlsx")):
            if ai_file.name.startswith("~$"):
                continue

            # マッチする正解ファイルを検索
            matched_correct = None
            for stem, correct_file in correct_files.items():
                # 物件名を抽出して比較
                ai_name = self._extract_property_name(ai_file.stem)
                correct_name = self._extract_property_name(stem)
                if ai_name and correct_name and (
                    ai_name in correct_name or correct_name in ai_name
                ):
                    matched_correct = correct_file
                    break

            if matched_correct:
                try:
                    report = self.compare(ai_file, matched_correct)
                    reports.append(report)
                except Exception as e:
                    logger.warning(f"比較エラー {ai_file.name}: {e}")

        return reports

    # ----------------------------------------------------------
    # ヘッダー比較
    # ----------------------------------------------------------

    def _compare_header(self, ws_ai, ws_correct) -> list[CellDiff]:
        """☆入力ヘッダー（rows 5-9）を比較"""
        diffs = []
        header_fields = [
            ("C5", "物件名", "major"),
            ("C6", "住所", "minor"),
            ("C7", "解錠番号", "minor"),
            ("C8", "分電盤", "minor"),
            ("C9", "特記事項", "minor"),
        ]
        for cell_ref, field_name, severity in header_fields:
            ai_val = self._safe_str(ws_ai[cell_ref].value)
            correct_val = self._safe_str(ws_correct[cell_ref].value)
            if self._normalize(ai_val) != self._normalize(correct_val):
                # 両方空は差分としない
                if not ai_val and not correct_val:
                    continue
                diffs.append(CellDiff(
                    sheet="☆入力",
                    cell=cell_ref,
                    field_name=field_name,
                    ai_value=ai_val,
                    correct_value=correct_val,
                    severity=severity,
                ))
        return diffs

    # ----------------------------------------------------------
    # 器具行比較
    # ----------------------------------------------------------

    def _compare_fixtures(self, ws_ai, ws_correct) -> list[FixtureRowDiff]:
        """☆入力の器具行（rows 16-45）を比較

        行の対応付けはrow_label（A, B, C...）ではなく、
        D列（照明種別）の内容で最もよいマッチを探す。
        ただし基本は同じ行番号同士を比較する。
        """
        start_row = INPUT_SHEET["data_start_row"]  # 16
        end_row = INPUT_SHEET["data_end_row"]       # 45
        results = []

        for i, row_num in enumerate(range(start_row, end_row + 1)):
            label = ROW_LABELS[i] if i < len(ROW_LABELS) else f"R{i}"

            ai_type = self._safe_str(ws_ai[f"D{row_num}"].value)
            correct_type = self._safe_str(ws_correct[f"D{row_num}"].value)

            ai_has = bool(ai_type)
            correct_has = bool(correct_type)

            if not ai_has and not correct_has:
                continue  # 両方空行

            row_diff = FixtureRowDiff(
                row_number=row_num,
                row_label=label,
                fixture_type_ai=ai_type,
                fixture_type_correct=correct_type,
            )

            if not ai_has and correct_has:
                row_diff.status = "missing_in_ai"
                results.append(row_diff)
                continue

            if ai_has and not correct_has:
                row_diff.status = "extra_in_ai"
                results.append(row_diff)
                continue

            # 両方データあり → セルレベルで比較
            cell_diffs = self._compare_fixture_row(ws_ai, ws_correct, row_num)
            row_diff.diffs = cell_diffs
            row_diff.status = "match" if not cell_diffs else "modified"
            results.append(row_diff)

        return results

    def _compare_fixture_row(
        self, ws_ai, ws_correct, row: int,
    ) -> list[CellDiff]:
        """1器具行の全フィールドを比較"""
        diffs = []
        fields = [
            ("C", "場所(C列)", "major"),
            ("D", "照明種別(D列)", "critical"),
            ("E", "現調備考(E列)", "minor"),
            ("F", "工事備考(F列)", "minor"),
            ("G", "LED選定(G列)", "critical"),
            ("I", "一日点灯(I列)", "minor"),
            ("K", "消費電力(K列)", "major"),
            ("L", "電球数(L列)", "critical"),
            ("AE", "工事単価(AE列)", "major"),
        ]

        for col, field_name, severity in fields:
            ai_val = self._safe_str(ws_ai[f"{col}{row}"].value)
            correct_val = self._safe_str(ws_correct[f"{col}{row}"].value)

            if self._normalize(ai_val) != self._normalize(correct_val):
                if not ai_val and not correct_val:
                    continue
                diffs.append(CellDiff(
                    sheet="☆入力",
                    cell=f"{col}{row}",
                    field_name=field_name,
                    ai_value=ai_val,
                    correct_value=correct_val,
                    severity=severity,
                ))

        # 階別数量（M-V列）
        floor_start = 13  # M列
        for floor_idx in range(10):
            col_num = floor_start + floor_idx
            col_letter = get_column_letter(col_num)
            ai_val = self._safe_str(
                ws_ai.cell(row=row, column=col_num).value)
            correct_val = self._safe_str(
                ws_correct.cell(row=row, column=col_num).value)
            if self._normalize(ai_val) != self._normalize(correct_val):
                if not ai_val and not correct_val:
                    continue
                # "0"と空は同等と見なす
                if self._is_zero(ai_val) and self._is_zero(correct_val):
                    continue
                diffs.append(CellDiff(
                    sheet="☆入力",
                    cell=f"{col_letter}{row}",
                    field_name=f"{floor_idx+1}F数量({col_letter}列)",
                    ai_value=ai_val,
                    correct_value=correct_val,
                    severity="major",
                ))

        return diffs

    # ----------------------------------------------------------
    # 除外行比較
    # ----------------------------------------------------------

    def _compare_excluded(self, ws_ai, ws_correct) -> list[FixtureRowDiff]:
        """☆入力の除外行（rows 49+）を比較"""
        start_row = INPUT_SHEET["excluded_start_row"]  # 49
        results = []

        for i in range(12):  # 最大12行
            row_num = start_row + i
            label = str(i + 1)

            ai_type = self._safe_str(ws_ai[f"D{row_num}"].value)
            correct_type = self._safe_str(ws_correct[f"D{row_num}"].value)

            ai_has = bool(ai_type)
            correct_has = bool(correct_type)

            if not ai_has and not correct_has:
                continue

            row_diff = FixtureRowDiff(
                row_number=row_num,
                row_label=label,
                fixture_type_ai=ai_type,
                fixture_type_correct=correct_type,
            )

            if not ai_has and correct_has:
                row_diff.status = "missing_in_ai"
                results.append(row_diff)
                continue

            if ai_has and not correct_has:
                row_diff.status = "extra_in_ai"
                results.append(row_diff)
                continue

            # セル比較
            cell_diffs = self._compare_excluded_row(
                ws_ai, ws_correct, row_num)
            row_diff.diffs = cell_diffs
            row_diff.status = "match" if not cell_diffs else "modified"
            results.append(row_diff)

        return results

    def _compare_excluded_row(
        self, ws_ai, ws_correct, row: int,
    ) -> list[CellDiff]:
        """1除外行のフィールド比較"""
        diffs = []
        fields = [
            ("C", "場所(C列)", "major"),
            ("D", "照明種別(D列)", "critical"),
            ("E", "現調備考(E列)", "minor"),
            ("L", "電球数(L列)", "major"),
            ("W", "除外理由(W列)", "major"),
        ]
        for col, field_name, severity in fields:
            ai_val = self._safe_str(ws_ai[f"{col}{row}"].value)
            correct_val = self._safe_str(ws_correct[f"{col}{row}"].value)
            if self._normalize(ai_val) != self._normalize(correct_val):
                if not ai_val and not correct_val:
                    continue
                diffs.append(CellDiff(
                    sheet="☆入力",
                    cell=f"{col}{row}",
                    field_name=field_name,
                    ai_value=ai_val,
                    correct_value=correct_val,
                    severity=severity,
                ))
        return diffs

    # ----------------------------------------------------------
    # 選定シート比較
    # ----------------------------------------------------------

    def _compare_selection(self, ws_ai, ws_correct) -> list[SelectionRowDiff]:
        """選定シートを比較"""
        start_row = SELECTION_SHEET["data_start_row"]  # 3
        results = []

        for row_num in range(start_row, start_row + 30):
            ai_key = self._safe_str(ws_ai[f"C{row_num}"].value)
            correct_key = self._safe_str(ws_correct[f"C{row_num}"].value)

            ai_has = bool(ai_key) and "***" not in ai_key
            correct_has = bool(correct_key) and "***" not in correct_key

            if not ai_has and not correct_has:
                continue

            row_diff = SelectionRowDiff(
                row_number=row_num,
                product_key_ai=ai_key,
                product_key_correct=correct_key,
            )

            if not ai_has and correct_has:
                row_diff.status = "missing_in_ai"
                results.append(row_diff)
                continue

            if ai_has and not correct_has:
                row_diff.status = "extra_in_ai"
                results.append(row_diff)
                continue

            # セル比較
            cell_diffs = self._compare_selection_row(
                ws_ai, ws_correct, row_num)
            row_diff.diffs = cell_diffs
            row_diff.status = "match" if not cell_diffs else "modified"
            results.append(row_diff)

        return results

    def _compare_selection_row(
        self, ws_ai, ws_correct, row: int,
    ) -> list[CellDiff]:
        """選定シート1行の比較"""
        diffs = []
        fields = [
            ("C", "商品名(C列)", "critical"),
            ("D", "照明色(D列)", "minor"),
            ("E", "器具色(E列)", "minor"),
            ("F", "器具サイズ(F列)", "minor"),
            ("G", "消費電力(G列)", "major"),
            ("H", "全光束(H列)", "minor"),
            ("I", "合算定価(I列)", "major"),
            ("J", "合算仕入(J列)", "major"),
            ("K", "防滴(K列)", "major"),
            ("M", "メーカー(M列)", "minor"),
            ("O", "器具型番(O列)", "major"),
            ("AG", "交換方法(AG列)", "major"),
        ]
        for col, field_name, severity in fields:
            ai_val = self._safe_str(ws_ai[f"{col}{row}"].value)
            correct_val = self._safe_str(ws_correct[f"{col}{row}"].value)
            if self._normalize(ai_val) != self._normalize(correct_val):
                if not ai_val and not correct_val:
                    continue
                # 数値の比較: 丸め誤差を許容
                if self._numeric_equal(ai_val, correct_val):
                    continue
                diffs.append(CellDiff(
                    sheet="選定",
                    cell=f"{col}{row}",
                    field_name=field_name,
                    ai_value=ai_val,
                    correct_value=correct_val,
                    severity=severity,
                ))
        return diffs

    # ----------------------------------------------------------
    # ユーティリティ
    # ----------------------------------------------------------

    def _find_sheet(self, wb, keyword: str):
        for sn in wb.sheetnames:
            if keyword in sn:
                return wb[sn]
        return None

    def _safe_str(self, value) -> str:
        if value is None:
            return ""
        s = str(value).strip()
        if s in ("0", "None", "0.0"):
            return ""
        return s

    def _normalize(self, text: str) -> str:
        if not text:
            return ""
        return unicodedata.normalize("NFKC", text.strip())

    def _is_zero(self, val: str) -> bool:
        if not val:
            return True
        try:
            return float(val) == 0
        except (ValueError, TypeError):
            return False

    def _numeric_equal(self, a: str, b: str, tolerance: float = 0.01) -> bool:
        try:
            return abs(float(a) - float(b)) < tolerance
        except (ValueError, TypeError):
            return False

    def _extract_property_name(self, filename: str) -> str:
        """ファイル名から物件名を抽出

        AI出力: 【LED導入ｼﾐｭﾚｰｼｮﾝ】コーポ八の通り → コーポ八の通り
        正解:   コーポ八の通り【LED導入ｼﾐｭﾚｰｼｮﾝ】 → コーポ八の通り
        """
        name = filename
        for tag in ["【LED導入ｼﾐｭﾚｰｼｮﾝ】", "【LED導入シミュレーション】"]:
            if tag in name:
                parts = name.split(tag)
                # タグの前後どちらかに物件名がある
                before = parts[0].strip()
                after = parts[-1].strip() if len(parts) > 1 else ""
                # 長い方を物件名とする
                name = after if len(after) >= len(before) else before
                break
        return name.strip()


def get_column_letter(col_idx: int) -> str:
    """列番号→列文字（1-based）"""
    result = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


# ============================================================
# CLI テスト
# ============================================================

if __name__ == "__main__":
    import sys
    import io
    sys.stdout = io.TextIOWrapper(
        sys.stdout.buffer, encoding="utf-8", errors="replace")

    logging.basicConfig(
        level=logging.INFO,
        format="%(levelname)s: %(message)s",
    )

    base = Path(__file__).parent.parent
    ai_dir = base / "output"
    correct_dir = base / "正しい見積り"
    feedback_dir = base / "feedback"

    comparator = FeedbackComparator()
    reports = comparator.compare_folder(ai_dir, correct_dir)

    for report in reports:
        print(report.print_summary())
        print()

        # JSON保存
        json_name = f"feedback_{report.property_name}.json"
        report.save_json(feedback_dir / json_name)

    if not reports:
        print("比較対象のファイルペアが見つかりませんでした。")
        print(f"  AI出力: {ai_dir}")
        print(f"  正解:   {correct_dir}")
