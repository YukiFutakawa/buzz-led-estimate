"""ラインナップ表（LED商品カタログ）の読み込みとインデックス化"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Optional

import openpyxl

from models import LEDProduct, FixtureCategory, LineupFile

logger = logging.getLogger(__name__)

# ラインナップ表で読み込み対象のシート名（計算式・依頼書等は除外）
LINEUP_SHEETS: dict[str, list[str]] = {
    LineupFile.FLUORESCENT: [
        "40形蛍光灯 ",
        "20形蛍光灯 ",
        "非常灯40形 ",
        "非常灯20形 ",
        "その他非常灯",
        "屋外ﾌﾞﾗｹｯﾄ",
        "直管形LED器具",
    ],
    LineupFile.OTHER: [
        "天井・壁面 ",
        "ﾎﾟｰﾁ・支柱",
        "DL※高出力",
        "ﾀﾞｳﾝﾗｲﾄ",
        "ｽﾎﾟｯﾄﾗｲﾄ",
        "LED球",
        "外部・ﾊﾞｲﾊﾟｽ",
        "EEスイッチ他",
        "バイパスコーウェル",
        "丸・四角(大)",      # ★Fix52: 埋込スクエアライト/ラウンドベースライト用
    ],
    LineupFile.MANIAC: [
        "誘導灯 各社",
        "ﾌｯﾄﾗｲﾄ",
        "庭園灯",
        "表札・ﾎﾟｰﾁ（250lm以下）",
        "ｱｸｾﾝﾄﾗｲﾄ",
        "人感ｾﾝｻ",
        "ﾎﾟｰﾁﾗｲﾄ",
        "筒型ブラ",
        "屋内",
        "門柱灯",
        "防犯灯",
        "ポール灯",
        "投光器・高天井",
        "ｱｰﾑｽﾎﾟｯﾄ",
    ],
}

# 列インデックス（0始まり）→ LEDProduct属性マッピング
# ラインナップ表のA-AJ列に対応
COL_MAP = {
    # A=0: 写真① (画像、スキップ)
    # B=1: 写真② (画像、スキップ)
    2: "name",                    # C: 名称/電球種別
    3: "lighting_color",          # D: 照明色
    4: "fixture_color",           # E: 器具色
    5: "fixture_size",            # F: 器具サイズ
    6: "power_w",                 # G: 消費電力
    7: "lumens",                  # H: 全光束
    8: "list_price_total",        # I: 合算定価
    9: "purchase_price_total",    # J: 合算仕入
    10: "_waterproof",            # K: 防滴（〇/✕→bool変換）
    11: "bulb_type",              # L: 電球種別
    12: "manufacturer",           # M: メーカー
    13: "watt_equivalent",        # N: W相当
    14: "model_number",           # O: 型番
    15: "model_price",            # P: 定価
    16: "model_purchase",         # Q: 仕入れ
    17: "model_number_2",         # R: 型番②
    18: "model_price_2",          # S: 定価②
    19: "model_purchase_2",       # T: 仕入②
    20: "model_number_3",         # U: 型番③
    21: "model_price_3",          # V: 定価③
    22: "model_purchase_3",       # W: 仕入③
    23: "model_number_4",         # X: 型番④
    24: "model_price_4",          # Y: 定価④
    25: "model_purchase_4",       # Z: 仕入④
    26: "power_detail",           # AA: 消費電力
    27: "lumens_detail",          # AB: 全光束
    28: "material",               # AC: 器具素材
    29: "color_options",          # AD: 器具色選択肢
    30: "lighting_color_options", # AE: 照明色選択肢
    31: "lifespan",               # AF: 定格寿命
    32: "replacement_method",     # AG: 交換方法
    33: "socket",                 # AH: 口金
    # AI=34: HP
    # AJ=35: 備考
}


def _safe_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _safe_int(val) -> int:
    if val is None:
        return 0
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return 0


def _safe_float(val) -> float:
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _is_waterproof(val) -> bool:
    s = _safe_str(val)
    return "〇" in s or "○" in s or "◯" in s


def _is_data_row(row_values: list) -> bool:
    """データ行かどうか判定（ヘッダー行・空行・セクション見出し行をスキップ）"""
    # C列（名称）に値があり、かつG列（消費電力）またはO列（型番）にも値があれば有効行
    name = row_values[2] if len(row_values) > 2 else None
    model = row_values[14] if len(row_values) > 14 else None
    power = row_values[6] if len(row_values) > 6 else None

    if not name:
        return False
    name_str = str(name).strip()
    if not name_str:
        return False
    # アスタリスク行（ドロップダウン用データ）をスキップ
    if name_str.startswith("*") or name_str.startswith("＊"):
        return False

    # 型番か消費電力のいずれかがあれば有効
    return model is not None or power is not None


def _parse_row(row_values: list, source_file: str, source_sheet: str,
               row_num: int) -> Optional[LEDProduct]:
    """1行分のデータをLEDProductに変換"""
    if not _is_data_row(row_values):
        return None

    product = LEDProduct(
        source_file=source_file,
        source_sheet=source_sheet.strip(),
        source_row=row_num,
    )

    for col_idx, attr_name in COL_MAP.items():
        if col_idx >= len(row_values):
            continue
        val = row_values[col_idx]

        if attr_name == "_waterproof":
            product.is_waterproof = _is_waterproof(val)
        elif attr_name in ("power_w", "power_detail"):
            setattr(product, attr_name, _safe_float(val))
        elif attr_name in (
            "list_price_total", "purchase_price_total",
            "model_price", "model_purchase",
            "model_price_2", "model_purchase_2",
            "model_price_3", "model_purchase_3",
            "model_price_4", "model_purchase_4",
        ):
            setattr(product, attr_name, _safe_int(val))
        else:
            setattr(product, attr_name, _safe_str(val))

    # HP列とbemerkungen列
    if len(row_values) > 34:
        product.hp_link = _safe_str(row_values[34])
    if len(row_values) > 35:
        product.notes = _safe_str(row_values[35])

    return product


class LineupIndex:
    """全ラインナップ表の検索可能インデックス"""

    def __init__(self):
        self.products: list[LEDProduct] = []
        self.by_sheet: dict[str, list[LEDProduct]] = {}
        self.by_manufacturer: dict[str, list[LEDProduct]] = {}

    def load_all(self, lineup_dir: Path) -> None:
        """ラインナップ表ディレクトリから全ファイルを読み込み"""
        for file_enum, sheets in LINEUP_SHEETS.items():
            filepath = lineup_dir / file_enum
            if not filepath.exists():
                logger.warning(f"ラインナップファイルが見つかりません: {filepath}")
                continue

            logger.info(f"読み込み中: {filepath.name}")
            self._load_file(filepath, file_enum, sheets)

        logger.info(
            f"ラインナップ読み込み完了: {len(self.products)}商品, "
            f"{len(self.by_sheet)}カテゴリ"
        )

    def _load_file(self, filepath: Path, file_name: str,
                   target_sheets: list[str]) -> None:
        """1つのラインナップファイルを読み込み"""
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)

        for sheet_name in wb.sheetnames:
            # 対象シートのみ処理（前後の空白を考慮して比較）
            matched_target = None
            for target in target_sheets:
                if sheet_name.strip() == target.strip():
                    matched_target = target
                    break
            if matched_target is None:
                continue

            ws = wb[sheet_name]
            sheet_key = sheet_name.strip()
            if sheet_key not in self.by_sheet:
                self.by_sheet[sheet_key] = []

            count = 0
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                row_values = list(row)
                product = _parse_row(row_values, file_name, sheet_name, row_num)
                if product is None:
                    continue

                self.products.append(product)
                self.by_sheet[sheet_key].append(product)

                # メーカー別インデックス
                if product.manufacturer:
                    mfr = product.manufacturer
                    if mfr not in self.by_manufacturer:
                        self.by_manufacturer[mfr] = []
                    self.by_manufacturer[mfr].append(product)

                count += 1

            logger.info(f"  {sheet_key}: {count}商品")

        wb.close()

    def search(
        self,
        sheet_name: Optional[str] = None,
        manufacturer: Optional[str] = None,
        waterproof: Optional[bool] = None,
        lighting_color: Optional[str] = None,
        max_power_w: Optional[float] = None,
        keyword: Optional[str] = None,
    ) -> list[LEDProduct]:
        """条件に合うLED商品を検索"""
        candidates = self.products

        if sheet_name:
            candidates = self.by_sheet.get(sheet_name.strip(), [])

        results = []
        for p in candidates:
            if manufacturer and manufacturer not in p.manufacturer:
                continue
            if waterproof is not None and p.is_waterproof != waterproof:
                continue
            if lighting_color:
                if (lighting_color not in p.lighting_color and
                        lighting_color not in p.lighting_color_options):
                    continue
            if max_power_w is not None and p.power_w > max_power_w:
                continue
            if keyword:
                kw = keyword.lower()
                searchable = (
                    p.name + p.bulb_type + p.model_number +
                    p.watt_equivalent + p.notes
                ).lower()
                if kw not in searchable:
                    continue
            results.append(p)

        return results

    def get_categories(self) -> list[str]:
        """利用可能なカテゴリ（シート名）一覧"""
        return sorted(self.by_sheet.keys())

    def get_stats(self) -> dict[str, int]:
        """カテゴリ別商品数"""
        return {k: len(v) for k, v in self.by_sheet.items()}
