# -*- coding: utf-8 -*-
"""正しい見積りExcelからパターンを自動抽出するインポーター

人間の専門家が作成した正しい見積りファイルを読み込み、
(既存器具 → LED選定) のマッピングパターンを構造化データとして抽出する。
"""

from __future__ import annotations

import logging
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.utils import column_index_from_string

from models import INPUT_SHEET, SELECTION_SHEET

logger = logging.getLogger(__name__)


# ============================================================
# データクラス
# ============================================================

@dataclass
class CorrectFixtureMapping:
    """☆入力 rows 16-45 の1行: 既存器具→LED選定のマッピング"""
    row_number: int              # Excel行番号
    row_label: str               # A列の行ラベル (A,B,C...)
    property_name: str           # B列: 物件名
    location: str                # C列: 試算エリア
    fixture_type: str            # D列: 照明種別
    size_memo: str               # E列: 現調備考（サイズ情報）
    construction_memo: str       # F列: 工事備考（電球種別等）
    led_selection: str           # G列: 器具分類②（★正解のLED選定）
    monthly_hours: float         # H列: 月間点灯
    daily_hours: float           # I列: 一日点灯
    operating_days: float        # J列: 稼働日数
    power_w: float               # K列: 消費電力
    quantity: int                # L列: 電球数
    floor_quantities: dict       # M-V列: {1: n, 2: m, ...}
    construction_price: float    # AE列: 工事単価
    profit_ratio: float          # AN列: 利益率


@dataclass
class CorrectExcludedFixture:
    """☆入力 rows 49+ の除外行"""
    row_number: int              # Excel行番号
    property_name: str           # B列
    location: str                # C列: 試算エリア
    fixture_type: str            # D列: 照明種別
    size_memo: str               # E列: 現調備考
    quantity: int                # L列: 電球数
    floor_quantities: dict       # M-V列
    exclusion_reason: str        # W列: 除外理由


@dataclass
class CorrectProductSpec:
    """選定シート rows 3+ の商品仕様"""
    row_number: int              # Excel行番号
    product_name: str            # C列（☆入力G列と一致するキー）
    lighting_color: str          # D列: 照明色
    fixture_color: str           # E列: 器具色
    fixture_size: str            # F列: 器具サイズ
    power_w: float               # G列: 消費電力
    lumens: float                # H列: 全光束
    list_price: float            # I列: 合算定価
    purchase_price: float        # J列: 合算仕入
    is_waterproof: bool          # K列: 防滴 (〇/✕)
    bulb_type: str               # L列: 電球種別
    manufacturer: str            # M列: メーカー
    watt_equivalent: str         # N列: W相当
    model_number: str            # O列: 器具型番
    model_price: float           # P列: 定価
    replacement_method: str      # AG列: 交換方法


@dataclass
class CorrectEstimate:
    """1つの正しい見積りファイルの全データ"""
    file_path: Path
    property_name: str
    address: str
    fixture_mappings: list[CorrectFixtureMapping]
    excluded_fixtures: list[CorrectExcludedFixture]
    product_specs: list[CorrectProductSpec]

    def summary(self) -> str:
        """概要文字列"""
        return (
            f"{self.property_name}: "
            f"器具{len(self.fixture_mappings)}件, "
            f"除外{len(self.excluded_fixtures)}件, "
            f"選定商品{len(self.product_specs)}件"
        )


# ============================================================
# インポーター本体
# ============================================================

class CorrectEstimateImporter:
    """正しい見積りExcelからパターンを抽出するインポーター"""

    def import_file(self, file_path: Path) -> CorrectEstimate:
        """1つの正しい見積りファイルを読み込み"""
        logger.info(f"正解ファイル読込: {file_path.name}")

        wb = openpyxl.load_workbook(file_path, data_only=True)

        ws_input = self._find_sheet(wb, '入力')
        ws_selection = self._find_sheet(wb, '選定')

        if not ws_input:
            raise ValueError(f"☆入力シートが見つかりません: {file_path.name}")

        # 物件情報
        property_name = self._safe_str(ws_input['C5'].value)
        address = self._safe_str(ws_input['C6'].value)

        # 器具マッピング
        mappings = self._read_fixture_mappings(ws_input)
        logger.info(f"  器具マッピング: {len(mappings)}件")

        # 除外器具
        excluded = self._read_excluded_fixtures(ws_input)
        logger.info(f"  除外器具: {len(excluded)}件")

        # 選定商品仕様
        products = []
        if ws_selection:
            products = self._read_product_specs(ws_selection)
            logger.info(f"  選定商品: {len(products)}件")

        wb.close()

        return CorrectEstimate(
            file_path=file_path,
            property_name=property_name,
            address=address,
            fixture_mappings=mappings,
            excluded_fixtures=excluded,
            product_specs=products,
        )

    def import_folder(self, folder_path: Path) -> list[CorrectEstimate]:
        """フォルダ内の全正しい見積りExcelを読み込み"""
        results = []
        for xlsx in sorted(folder_path.glob('*.xlsx')):
            if xlsx.name.startswith('~$'):
                continue  # Excelの一時ファイル
            try:
                result = self.import_file(xlsx)
                results.append(result)
                logger.info(f"  → {result.summary()}")
            except Exception as e:
                logger.warning(f"  読込失敗: {xlsx.name}: {e}")
        return results

    # ----------------------------------------------------------
    # 内部メソッド
    # ----------------------------------------------------------

    def _find_sheet(self, wb, keyword: str):
        """シート名にkeywordを含むシートを検索"""
        for sn in wb.sheetnames:
            if keyword in sn:
                return wb[sn]
        return None

    def _normalize(self, text: str) -> str:
        """テキストをNFKC正規化（全角半角統一）"""
        if text is None:
            return ""
        s = str(text).strip()
        return unicodedata.normalize('NFKC', s)

    def _safe_str(self, value) -> str:
        """セル値を安全に文字列化"""
        if value is None:
            return ""
        s = str(value).strip()
        if s in ('0', 'None'):
            return ""
        return s

    def _safe_float(self, value) -> float:
        """セル値を安全にfloat化"""
        if value is None:
            return 0.0
        try:
            return float(value)
        except (ValueError, TypeError):
            return 0.0

    def _safe_int(self, value) -> int:
        """セル値を安全にint化"""
        if value is None:
            return 0
        try:
            return int(float(value))
        except (ValueError, TypeError):
            return 0

    def _col_idx(self, col_letter: str) -> int:
        """列文字→列番号（1-based）"""
        return column_index_from_string(col_letter)

    def _read_fixture_mappings(self, ws) -> list[CorrectFixtureMapping]:
        """☆入力シートの器具行（rows 16-45）を読み取り"""
        mappings = []
        start_row = INPUT_SHEET["data_start_row"]  # 16
        end_row = INPUT_SHEET["data_end_row"]       # 45

        # 行ラベル（A, B, C...）
        row_labels = [
            "A", "B", "C", "D", "E", "F", "G", "H", "I", "J",
            "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T",
            "U", "V", "W", "X", "Y", "Z", "AA", "AB", "AC", "AD",
        ]

        for i, row_num in enumerate(range(start_row, end_row + 1)):
            # D列（照明種別）が空なら空行 → スキップ
            fixture_type = self._safe_str(ws[f'D{row_num}'].value)
            if not fixture_type:
                continue

            # 階別数量 (M-V列 = 1F-10F)
            floor_start_col = self._col_idx(INPUT_SHEET["col_floor_start"])  # M=13
            floor_quantities = {}
            for floor_idx in range(10):
                col = floor_start_col + floor_idx
                val = self._safe_int(ws.cell(row=row_num, column=col).value)
                if val > 0:
                    floor_quantities[floor_idx + 1] = val

            # AN列 = 利益率（列番号40）
            profit_ratio = self._safe_float(ws[f'AN{row_num}'].value)

            mapping = CorrectFixtureMapping(
                row_number=row_num,
                row_label=row_labels[i] if i < len(row_labels) else f"R{i}",
                property_name=self._safe_str(ws[f'B{row_num}'].value),
                location=self._safe_str(ws[f'C{row_num}'].value),
                fixture_type=fixture_type,
                size_memo=self._safe_str(ws[f'E{row_num}'].value),
                construction_memo=self._safe_str(ws[f'F{row_num}'].value),
                led_selection=self._safe_str(ws[f'G{row_num}'].value),
                monthly_hours=self._safe_float(ws[f'H{row_num}'].value),
                daily_hours=self._safe_float(ws[f'I{row_num}'].value),
                operating_days=self._safe_float(ws[f'J{row_num}'].value),
                power_w=self._safe_float(ws[f'K{row_num}'].value),
                quantity=self._safe_int(ws[f'L{row_num}'].value),
                floor_quantities=floor_quantities,
                construction_price=self._safe_float(ws[f'AE{row_num}'].value),
                profit_ratio=profit_ratio,
            )
            mappings.append(mapping)

        return mappings

    def _read_excluded_fixtures(self, ws) -> list[CorrectExcludedFixture]:
        """☆入力シートの除外行（rows 49+）を読み取り

        構造:
          Row 47: 【LED済み箇所】ヘッダー
          Row 48: 列ヘッダー (B=物件名, C=試算エリア, D=照明種別, ...)
          Row 49+: データ行 (A=番号, B=物件名, C=場所, D=器具種別, W=除外理由)
        """
        excluded = []
        start_row = INPUT_SHEET["excluded_start_row"]  # 49

        for row_num in range(start_row, start_row + 12):  # 最大12行
            # B列（物件名）とD列（照明種別）の両方が空なら終了
            prop = self._safe_str(ws[f'B{row_num}'].value)
            fixture_type = self._safe_str(ws[f'D{row_num}'].value)
            if not prop and not fixture_type:
                continue

            # 階別数量 (M-V列)
            floor_start_col = self._col_idx(INPUT_SHEET["col_floor_start"])
            floor_quantities = {}
            for floor_idx in range(10):
                col = floor_start_col + floor_idx
                val = self._safe_int(ws.cell(row=row_num, column=col).value)
                if val > 0:
                    floor_quantities[floor_idx + 1] = val

            excl = CorrectExcludedFixture(
                row_number=row_num,
                property_name=prop,
                location=self._safe_str(ws[f'C{row_num}'].value),
                fixture_type=fixture_type,
                size_memo=self._safe_str(ws[f'E{row_num}'].value),
                quantity=self._safe_int(ws[f'L{row_num}'].value),
                floor_quantities=floor_quantities,
                exclusion_reason=self._safe_str(ws[f'W{row_num}'].value),
            )
            excluded.append(excl)

        return excluded

    def _read_product_specs(self, ws) -> list[CorrectProductSpec]:
        """選定シートの商品仕様（rows 3+）を読み取り"""
        products = []
        start_row = SELECTION_SHEET["data_start_row"]  # 3

        for row_num in range(start_row, start_row + 20):  # 最大20商品
            # C列（商品名）が空 or *** パターンで終了
            product_name = self._safe_str(ws[f'C{row_num}'].value)
            if not product_name or '***' in product_name:
                break

            # K列: 防滴判定
            waterproof_val = self._safe_str(ws[f'K{row_num}'].value)
            is_waterproof = waterproof_val in ('〇', '○', '◯', 'O')

            spec = CorrectProductSpec(
                row_number=row_num,
                product_name=product_name,
                lighting_color=self._safe_str(ws[f'D{row_num}'].value),
                fixture_color=self._safe_str(ws[f'E{row_num}'].value),
                fixture_size=self._safe_str(ws[f'F{row_num}'].value),
                power_w=self._safe_float(ws[f'G{row_num}'].value),
                lumens=self._safe_float(ws[f'H{row_num}'].value),
                list_price=self._safe_float(ws[f'I{row_num}'].value),
                purchase_price=self._safe_float(ws[f'J{row_num}'].value),
                is_waterproof=is_waterproof,
                bulb_type=self._safe_str(ws[f'L{row_num}'].value),
                manufacturer=self._safe_str(ws[f'M{row_num}'].value),
                watt_equivalent=self._safe_str(ws[f'N{row_num}'].value),
                model_number=self._safe_str(ws[f'O{row_num}'].value),
                model_price=self._safe_float(ws[f'P{row_num}'].value),
                replacement_method=self._safe_str(ws[f'AG{row_num}'].value),
            )
            products.append(spec)

        return products


# ============================================================
# CLI テスト
# ============================================================

if __name__ == '__main__':
    import sys, io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

    logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')

    base = Path(__file__).parent.parent
    correct_dir = base / '正しい見積り'

    importer = CorrectEstimateImporter()
    estimates = importer.import_folder(correct_dir)

    for est in estimates:
        print(f"\n{'='*60}")
        print(f"物件: {est.property_name}")
        print(f"住所: {est.address}")
        print(f"ファイル: {est.file_path.name}")
        print(f"{'='*60}")

        print(f"\n--- 器具→LED選定マッピング ({len(est.fixture_mappings)}件) ---")
        for m in est.fixture_mappings:
            print(f"  [{m.row_label}] {m.fixture_type:25s} | {m.size_memo:12s} | "
                  f"{m.construction_memo:8s} → {m.led_selection}")
            print(f"       場所={m.location}, 数量={m.quantity}, "
                  f"階別={m.floor_quantities}, 工事単価={m.construction_price}")

        if est.excluded_fixtures:
            print(f"\n--- 除外器具 ({len(est.excluded_fixtures)}件) ---")
            for e in est.excluded_fixtures:
                print(f"  {e.fixture_type:20s} | 数量={e.quantity} | "
                      f"場所={e.location}")
                print(f"    理由: {e.exclusion_reason}")

        if est.product_specs:
            print(f"\n--- 選定商品 ({len(est.product_specs)}件) ---")
            for p in est.product_specs:
                wpr = '防滴〇' if p.is_waterproof else '防滴✕'
                print(f"  {p.product_name}")
                print(f"    {p.manufacturer} | {p.model_number} | "
                      f"¥{p.purchase_price:,.0f} | {wpr} | {p.replacement_method}")
